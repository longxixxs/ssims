from io import BytesIO
from urllib.parse import urlencode
from unittest.mock import patch

from django.contrib.auth.models import Group, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from .models import (
    AuditLog,
    SelectionHistory,
    StudentImportJob,
    TeacherClassAssignment,
    TeacherCourseAssignment,
    UserAccount,
    cl,
    course,
    depart,
    sc,
    student,
)


DEFAULT_PASSWORD = "Pass123456!"
DEFAULT_STUDENT_PASSWORD = "psw123456"
MANAGED_ROLES = ("admin", "teacher", "student")


class BaseTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        for role in MANAGED_ROLES:
            Group.objects.get_or_create(name=role)

    def create_user(self, username, role=None, password=DEFAULT_PASSWORD, status=None):
        user = User.objects.create_user(username=username, password=password, first_name=username)
        if role:
            user.groups.add(Group.objects.get(name=role))
        if status is None:
            status = UserAccount.STATUS_ACTIVE if role else UserAccount.STATUS_PENDING
        UserAccount.objects.create(user=user, status=status)
        if status == UserAccount.STATUS_DISABLED:
            user.is_active = False
            user.save(update_fields=["is_active"])
        return user

    def assign_teacher(self, teacher_user, classes=None, courses=None):
        classes = classes or []
        courses = courses or []
        for cls in classes:
            TeacherClassAssignment.objects.get_or_create(teacher=teacher_user, class_obj=cls)
        for crs in courses:
            TeacherCourseAssignment.objects.get_or_create(teacher=teacher_user, course_obj=crs)

    def create_school_base(self, dno="D001", classno="C001", cno="K01"):
        dep = depart.objects.create(dno=dno, dname=f"系{dno[-1]}", telephone="123456")
        cls = cl.objects.create(classno=classno, classname=f"班{classno[-1]}", dno=dep)
        crs = course.objects.create(cno=cno, cname=f"课程{cno}", credit=2, semester=1, type="crc")
        return dep, cls, crs

    def build_student_excel_file(self, rows, headers=None):
        if headers is None:
            headers = [
                "sno", "sname", "sex", "native", "age",
                "classno", "semester", "home", "telephone",
            ]
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        return SimpleUploadedFile(
            "students.xlsx",
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class AuthAndAccountStatusTests(BaseTestCase):
    def setUp(self):
        self.dep, self.cls, _ = self.create_school_base()
        self.admin = self.create_user("admin_auth", role="admin")

    def test_register_creates_pending_account_profile(self):
        resp = self.client.post(
            reverse("register"),
            {
                "username": "pending_user",
                "nickname": "待审核",
                "password1": "Pending123!",
                "password2": "Pending123!",
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse("login"))

        user_obj = User.objects.get(username="pending_user")
        self.assertEqual(user_obj.groups.count(), 0)
        self.assertEqual(user_obj.account_profile.status, UserAccount.STATUS_PENDING)

    def test_pending_account_cannot_login(self):
        self.create_user("pending_login", role=None, password="Pending123!", status=UserAccount.STATUS_PENDING)
        resp = self.client.post(
            reverse("login"),
            {"username": "pending_login", "password": "Pending123!"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse("login"))
        self.assertNotIn("_auth_user_id", self.client.session)

    def test_disabled_account_cannot_login(self):
        self.create_user("disabled_user", role="teacher", password="Teacher123!", status=UserAccount.STATUS_DISABLED)
        resp = self.client.post(
            reverse("login"),
            {"username": "disabled_user", "password": "Teacher123!"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertNotIn("_auth_user_id", self.client.session)

    def test_student_login_redirects_to_detail(self):
        stu_user = self.create_user("S2001", role="student", password="Student123!")
        stu = student.objects.create(sno="S2001", user=stu_user, sname="学生2001", sex="boy", classno=self.cls)
        resp = self.client.post(reverse("login"), {"username": "S2001", "password": "Student123!"})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse("student_detail", args=[stu.sno]))

    def test_teacher_login_redirects_dashboard(self):
        self.create_user("teacher_login", role="teacher", password="Teacher123!")
        resp = self.client.post(reverse("login"), {"username": "teacher_login", "password": "Teacher123!"})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse("dashboard"))

    def test_logout_requires_post(self):
        self.assertTrue(self.client.login(username="admin_auth", password=DEFAULT_PASSWORD))
        self.assertEqual(self.client.get(reverse("logout")).status_code, 405)
        resp = self.client.post(reverse("logout"))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse("login"))

    def test_legacy_student_default_password_backfills_active_account(self):
        student.objects.create(sno="S2002", sname="学生2002", sex="girl", classno=self.cls)
        resp = self.client.post(reverse("login"), {"username": "S2002", "password": DEFAULT_STUDENT_PASSWORD})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse("student_detail", args=["S2002"]))
        user_obj = User.objects.get(username="S2002")
        self.assertEqual(user_obj.account_profile.status, UserAccount.STATUS_ACTIVE)


class UserApprovalAndAssignmentTests(BaseTestCase):
    def setUp(self):
        self.admin = self.create_user("admin_user", role="admin")
        self.dep, self.cls, self.crs = self.create_school_base(dno="D011", classno="C011", cno="K11")

    def test_user_list_is_admin_only(self):
        self.assertTrue(self.client.login(username="admin_user", password=DEFAULT_PASSWORD))
        self.assertEqual(self.client.get(reverse("user_list")).status_code, 200)
        self.client.logout()

        teacher = self.create_user("teacher_user", role="teacher")
        self.assertTrue(self.client.login(username="teacher_user", password=DEFAULT_PASSWORD))
        resp = self.client.get(reverse("user_list"))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, "/")

    def test_teacher_user_requires_assignment_on_create(self):
        self.assertTrue(self.client.login(username="admin_user", password=DEFAULT_PASSWORD))
        resp = self.client.post(
            reverse("user_add"),
            {
                "username": "teacher_no_scope",
                "nickname": "无归属教师",
                "password1": "Teacher123!",
                "password2": "Teacher123!",
                "groups": ["teacher"],
                "status": UserAccount.STATUS_ACTIVE,
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(User.objects.filter(username="teacher_no_scope").exists())

    def test_teacher_user_create_saves_status_and_assignments(self):
        self.assertTrue(self.client.login(username="admin_user", password=DEFAULT_PASSWORD))
        resp = self.client.post(
            reverse("user_add"),
            {
                "username": "teacher_scoped",
                "nickname": "负责教师",
                "password1": "Teacher123!",
                "password2": "Teacher123!",
                "groups": ["teacher"],
                "status": UserAccount.STATUS_ACTIVE,
                "teacher_classes": [self.cls.classno],
                "teacher_courses": [self.crs.cno],
            },
        )
        self.assertEqual(resp.status_code, 302)
        teacher_user = User.objects.get(username="teacher_scoped")
        self.assertEqual(teacher_user.account_profile.status, UserAccount.STATUS_ACTIVE)
        self.assertTrue(TeacherClassAssignment.objects.filter(teacher=teacher_user, class_obj=self.cls).exists())
        self.assertTrue(TeacherCourseAssignment.objects.filter(teacher=teacher_user, course_obj=self.crs).exists())
        self.assertTrue(AuditLog.objects.filter(action="create", model_name="User", object_id=str(teacher_user.id)).exists())

    def test_pending_approval_queue_lists_pending_users(self):
        pending = self.create_user("pending_review", role=None, status=UserAccount.STATUS_PENDING)
        active = self.create_user("active_user", role="teacher", status=UserAccount.STATUS_ACTIVE)
        self.assertTrue(self.client.login(username="admin_user", password=DEFAULT_PASSWORD))
        resp = self.client.get(reverse("pending_approval_list"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, pending.username)
        self.assertNotContains(resp, active.username)

    def test_pending_approval_queue_is_admin_only(self):
        teacher = self.create_user("teacher_denied", role="teacher")
        self.assertTrue(self.client.login(username="teacher_denied", password=DEFAULT_PASSWORD))
        resp = self.client.get(reverse("pending_approval_list"))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, "/")

    def test_admin_can_approve_registered_user_with_role_and_scope(self):
        resp = self.client.post(
            reverse("register"),
            {
                "username": "approval_target",
                "nickname": "待审批教师",
                "password1": "Pending123!",
                "password2": "Pending123!",
            },
        )
        self.assertEqual(resp.status_code, 302)
        pending_user = User.objects.get(username="approval_target")

        self.assertTrue(self.client.login(username="admin_user", password=DEFAULT_PASSWORD))
        edit_resp = self.client.post(
            reverse("user_edit", args=[pending_user.id]),
            {
                "nickname": "审批完成",
                "groups": ["teacher"],
                "status": UserAccount.STATUS_ACTIVE,
                "teacher_classes": [self.cls.classno],
            },
        )
        self.assertEqual(edit_resp.status_code, 302)
        pending_user.refresh_from_db()
        self.assertTrue(pending_user.groups.filter(name="teacher").exists())
        self.assertEqual(pending_user.account_profile.status, UserAccount.STATUS_ACTIVE)
        self.assertTrue(TeacherClassAssignment.objects.filter(teacher=pending_user, class_obj=self.cls).exists())

    def test_active_student_profile_blocks_role_change_to_non_student(self):
        stu_user = self.create_user("S3004", role="student")
        student.objects.create(sno="S3004", user=stu_user, sname="学生3004", sex="girl", classno=self.cls)
        self.assertTrue(self.client.login(username="admin_user", password=DEFAULT_PASSWORD))
        resp = self.client.post(
            reverse("user_edit", args=[stu_user.id]),
            {
                "nickname": "转教师",
                "groups": ["teacher"],
                "status": UserAccount.STATUS_ACTIVE,
                "teacher_classes": [self.cls.classno],
            },
        )
        self.assertEqual(resp.status_code, 200)
        stu_user.refresh_from_db()
        self.assertTrue(stu_user.groups.filter(name="student").exists())

    def test_create_student_user_uses_default_password(self):
        self.assertTrue(self.client.login(username="admin_user", password=DEFAULT_PASSWORD))
        resp = self.client.post(
            reverse("user_add"),
            {
                "username": "S3010",
                "nickname": "学生3010",
                "groups": ["student"],
                "status": UserAccount.STATUS_ACTIVE,
                "create_student": "on",
                "sname": "学生3010",
                "sex": "boy",
                "classno": self.cls.classno,
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.client.logout()
        self.assertTrue(self.client.login(username="S3010", password=DEFAULT_STUDENT_PASSWORD))


class StudentLifecycleAndImportTests(BaseTestCase):
    def setUp(self):
        self.admin = self.create_user("admin_stu", role="admin")
        self.teacher = self.create_user("teacher_stu", role="teacher")
        self.dep, self.cls, self.crs = self.create_school_base(dno="D021", classno="C021", cno="K21")
        self.assign_teacher(self.teacher, classes=[self.cls], courses=[self.crs])
        self.student_user = self.create_user("S2101", role="student")
        self.stu = student.objects.create(sno="S2101", user=self.student_user, sname="学生2101", sex="boy", classno=self.cls)

    def test_student_archive_disable_account(self):
        self.assertTrue(self.client.login(username="admin_stu", password=DEFAULT_PASSWORD))
        resp = self.client.post(reverse("student_delete", args=[self.stu.sno]), {"account_action": "disable"})
        self.assertEqual(resp.status_code, 302)
        self.stu.refresh_from_db()
        self.student_user.refresh_from_db()
        self.assertFalse(self.stu.is_active)
        self.assertEqual(self.student_user.account_profile.status, UserAccount.STATUS_DISABLED)
        self.assertFalse(self.student_user.groups.filter(name="student").exists())

    def test_student_archive_unlink_account(self):
        self.assertTrue(self.client.login(username="admin_stu", password=DEFAULT_PASSWORD))
        resp = self.client.post(reverse("student_delete", args=[self.stu.sno]), {"account_action": "unlink"})
        self.assertEqual(resp.status_code, 302)
        self.stu.refresh_from_db()
        self.student_user.refresh_from_db()
        self.assertIsNone(self.stu.user)
        self.assertEqual(self.student_user.account_profile.status, UserAccount.STATUS_PENDING)

    def test_student_add_can_restore_archived_record(self):
        self.stu.is_active = False
        self.stu.save(update_fields=["is_active"])
        self.assertTrue(self.client.login(username="admin_stu", password=DEFAULT_PASSWORD))
        resp = self.client.post(
            reverse("student_add"),
            {
                "sno": self.stu.sno,
                "sname": "恢复学生",
                "sex": "boy",
                "classno": self.cls.classno,
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.stu.refresh_from_db()
        self.assertTrue(self.stu.is_active)
        self.assertEqual(self.stu.sname, "恢复学生")

    def test_teacher_sees_only_assigned_students(self):
        dep2, cls2, _ = self.create_school_base(dno="D022", classno="C022", cno="K22")
        other_user = self.create_user("S2199", role="student")
        other_stu = student.objects.create(sno="S2199", user=other_user, sname="学生2199", sex="girl", classno=cls2)
        self.assertTrue(self.client.login(username="teacher_stu", password=DEFAULT_PASSWORD))
        resp = self.client.get(reverse("student_list"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, self.stu.sno)
        self.assertNotContains(resp, other_stu.sno)

    def test_import_preview_and_apply_create_mode(self):
        self.assertTrue(self.client.login(username="admin_stu", password=DEFAULT_PASSWORD))
        upload = self.build_student_excel_file(
            rows=[["S2110", "导入A", "boy", "北京", 20, self.cls.classno, 1, "海淀", "13800000001"]]
        )
        preview = self.client.post(reverse("student_import_excel"), {"file": upload, "mode": "create"})
        self.assertEqual(preview.status_code, 200)
        job = preview.context["job"]
        self.assertEqual(job.status, StudentImportJob.STATUS_PREVIEWED)
        self.assertEqual(len(job.preview_rows), 1)

        apply_resp = self.client.post(reverse("student_import_excel"), {"action": "apply", "job_id": job.id})
        self.assertEqual(apply_resp.status_code, 302)
        self.assertTrue(student.objects.filter(sno="S2110", is_active=True).exists())
        self.assertTrue(User.objects.filter(username="S2110").exists())

    def test_import_update_mode_updates_existing_student(self):
        self.assertTrue(self.client.login(username="admin_stu", password=DEFAULT_PASSWORD))
        upload = self.build_student_excel_file(
            rows=[[self.stu.sno, "已更新", "boy", "上海", 22, self.cls.classno, 2, "浦东", "13900000000"]]
        )
        preview = self.client.post(reverse("student_import_excel"), {"file": upload, "mode": "update"})
        self.assertEqual(preview.status_code, 200)
        job = preview.context["job"]
        self.assertEqual(job.preview_rows[0]["operation"], "update")

        self.client.post(reverse("student_import_excel"), {"action": "apply", "job_id": job.id})
        self.stu.refresh_from_db()
        self.assertEqual(self.stu.sname, "已更新")
        self.assertEqual(self.stu.semester, 2)

    def test_import_invalid_row_stays_in_error_list(self):
        self.assertTrue(self.client.login(username="admin_stu", password=DEFAULT_PASSWORD))
        upload = self.build_student_excel_file(
            rows=[["S2112", "非法学生", "invalid-sex", "北京", -1, self.cls.classno, 20, "海淀", "13800000003"]]
        )
        preview = self.client.post(reverse("student_import_excel"), {"file": upload, "mode": "create"})
        self.assertEqual(preview.status_code, 200)
        job = preview.context["job"]
        self.assertEqual(len(job.preview_rows), 0)
        self.assertEqual(len(job.error_rows), 1)

    def test_import_invalid_header_redirects(self):
        self.assertTrue(self.client.login(username="admin_stu", password=DEFAULT_PASSWORD))
        upload = self.build_student_excel_file(rows=[["S2113", "坏表头"]], headers=["bad1", "bad2"])
        resp = self.client.post(reverse("student_import_excel"), {"file": upload, "mode": "create"})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse("student_import_excel"))

    def test_teacher_export_only_contains_owned_students(self):
        dep2, cls2, _ = self.create_school_base(dno="D023", classno="C023", cno="K23")
        other_user = self.create_user("S2200", role="student")
        student.objects.create(sno="S2200", user=other_user, sname="学生2200", sex="boy", classno=cls2)
        self.assertTrue(self.client.login(username="teacher_stu", password=DEFAULT_PASSWORD))
        resp = self.client.get(reverse("student_export_excel"))
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        exported_snos = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertIn(self.stu.sno, exported_snos)
        self.assertNotIn("S2200", exported_snos)

    def test_student_list_age_sort_places_empty_values_last(self):
        student.objects.create(sno="S2102", sname="学生2102", sex="girl", age=18, classno=self.cls)
        student.objects.create(sno="S2103", sname="学生2103", sex="boy", age=20, classno=self.cls)

        self.assertTrue(self.client.login(username="admin_stu", password=DEFAULT_PASSWORD))
        resp = self.client.get(reverse("student_list"), {"order": "age", "direction": "asc"})

        self.assertEqual(resp.status_code, 200)
        ordered_snos = [item.sno for item in resp.context["students"]]
        self.assertEqual(ordered_snos[:3], ["S2102", "S2103", "S2101"])


class AcademicScopeAndWorkflowTests(BaseTestCase):
    def setUp(self):
        self.admin = self.create_user("admin_acad", role="admin")
        self.teacher = self.create_user("teacher_acad", role="teacher")
        self.dep, self.cls, self.crs = self.create_school_base(dno="D031", classno="C031", cno="K31")
        self.assign_teacher(self.teacher, classes=[self.cls], courses=[self.crs])
        self.stu_user = self.create_user("S3101", role="student")
        self.stu = student.objects.create(sno="S3101", user=self.stu_user, sname="学生3101", sex="boy", classno=self.cls)

    def test_depart_add_page_and_duplicate_post_render_without_missing_depart_context(self):
        self.assertTrue(self.client.login(username="admin_acad", password=DEFAULT_PASSWORD))

        get_resp = self.client.get(reverse("depart_add"))
        self.assertEqual(get_resp.status_code, 200)
        self.assertContains(get_resp, "新增系部")

        post_resp = self.client.post(reverse("depart_add"), {
            "dno": self.dep.dno,
            "dname": "重复系部",
            "telephone": "010-12345678",
        })
        self.assertEqual(post_resp.status_code, 200)
        self.assertContains(post_resp, "系部编号已存在")
        self.assertContains(post_resp, 'value="%s"' % self.dep.dno, html=False)

    def test_depart_class_course_archive_are_protected_by_active_dependencies(self):
        sc.objects.create(sno=self.stu, cno=self.crs)
        self.assertTrue(self.client.login(username="admin_acad", password=DEFAULT_PASSWORD))

        depart_resp = self.client.post(reverse("depart_delete", args=[self.dep.dno]))
        class_resp = self.client.post(reverse("class_delete", args=[self.cls.classno]))
        course_resp = self.client.post(reverse("course_delete", args=[self.crs.cno]))

        self.assertEqual(depart_resp.status_code, 302)
        self.assertEqual(class_resp.status_code, 302)
        self.assertEqual(course_resp.status_code, 302)
        self.dep.refresh_from_db()
        self.cls.refresh_from_db()
        self.crs.refresh_from_db()
        self.assertTrue(self.dep.is_active)
        self.assertTrue(self.cls.is_active)
        self.assertTrue(self.crs.is_active)

    def test_teacher_list_views_are_scoped(self):
        dep2, cls2, crs2 = self.create_school_base(dno="D032", classno="C032", cno="K32")
        self.assertTrue(self.client.login(username="teacher_acad", password=DEFAULT_PASSWORD))

        class_resp = self.client.get(reverse("class_list"))
        course_resp = self.client.get(reverse("course_list"))
        depart_resp = self.client.get(reverse("depart_list"))

        self.assertContains(class_resp, self.cls.classno)
        self.assertNotContains(class_resp, cls2.classno)
        self.assertContains(course_resp, self.crs.cno)
        self.assertNotContains(course_resp, crs2.cno)
        self.assertContains(depart_resp, self.dep.dno)
        self.assertNotContains(depart_resp, dep2.dno)

    def test_course_list_sort_links_preserve_encoded_filters(self):
        course.objects.create(cno="K98", cname="课&A=1", credit=3, semester=2, type="crc")
        self.assertTrue(self.client.login(username="admin_acad", password=DEFAULT_PASSWORD))

        resp = self.client.get(reverse("course_list"), {
            "cname": "课&A=1",
            "type": "crc",
        })

        expected = urlencode({"cname": "课&A=1", "type": "crc"})
        html_expected = expected.replace("&", "&amp;")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["sort_query_string"], expected)
        self.assertContains(resp, "?%s&amp;order=cno" % html_expected, html=False)
        self.assertContains(resp, "?%s&amp;order=cname" % html_expected, html=False)
        self.assertContains(resp, "?%s&amp;order=credit" % html_expected, html=False)

    def test_course_list_statistics_reflect_filtered_results(self):
        course.objects.create(cno="K97", cname="专业课A", credit=2, semester=2, type="spc")
        course.objects.create(cno="K96", cname="专业课B", credit=2, semester=2, type="spc")
        course.objects.create(cno="K95", cname="公共课A", credit=2, semester=1, type="crc")
        self.assertTrue(self.client.login(username="admin_acad", password=DEFAULT_PASSWORD))

        resp = self.client.get(reverse("course_list"), {
            "semester": "2",
            "type": "spc",
        })

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["current_semester_label"], "2")
        self.assertEqual(resp.context["current_type_label"], "专业课")
        self.assertEqual(resp.context["semester_count"], 1)
        self.assertEqual(resp.context["type_count"], 1)

    def test_row_link_lists_are_keyboard_accessible_and_blank_target_is_hardened(self):
        self.assertTrue(self.client.login(username="admin_acad", password=DEFAULT_PASSWORD))

        student_resp = self.client.get(reverse("student_list"))
        class_resp = self.client.get(reverse("class_list"))
        course_resp = self.client.get(reverse("course_list"))
        class_form_resp = self.client.get(reverse("class_add"))

        self.assertContains(student_resp, 'tabindex="0" role="link" aria-label="查看学生', html=False)
        self.assertContains(class_resp, 'tabindex="0" role="link" aria-label="查看班级', html=False)
        self.assertContains(course_resp, 'tabindex="0" role="link" aria-label="查看课程', html=False)
        self.assertContains(class_form_resp, 'target="_blank" rel="noopener noreferrer"', html=False)

    def test_teacher_can_only_select_assigned_course_for_assigned_student(self):
        dep2, cls2, crs2 = self.create_school_base(dno="D033", classno="C033", cno="K33")
        other_user = self.create_user("S3199", role="student")
        other_stu = student.objects.create(sno="S3199", user=other_user, sname="学生3199", sex="girl", classno=cls2)
        self.assertTrue(self.client.login(username="teacher_acad", password=DEFAULT_PASSWORD))

        allowed_get = self.client.get(reverse("select_course", args=[self.stu.sno]))
        denied_get = self.client.get(reverse("select_course", args=[other_stu.sno]))
        denied_post = self.client.post(reverse("select_course", args=[self.stu.sno]), {"cno": crs2.cno})

        self.assertEqual(allowed_get.status_code, 200)
        self.assertEqual(denied_get.status_code, 302)
        self.assertEqual(denied_post.status_code, 302)
        self.assertFalse(sc.objects.filter(sno=self.stu, cno=crs2).exists())

    def test_teacher_cannot_open_unassigned_course_students_page(self):
        _, _, crs2 = self.create_school_base(dno="D035", classno="C035", cno="K35")
        self.assertTrue(self.client.login(username="teacher_acad", password=DEFAULT_PASSWORD))
        resp = self.client.get(reverse("course_students", args=[crs2.cno]))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, "/")

    def test_drop_course_then_reselect_restores_same_record(self):
        record = sc.objects.create(sno=self.stu, cno=self.crs)
        self.assertTrue(self.client.login(username="teacher_acad", password=DEFAULT_PASSWORD))
        drop_resp = self.client.post(reverse("drop_course", args=[self.stu.sno, self.crs.cno]))
        self.assertEqual(drop_resp.status_code, 302)
        record.refresh_from_db()
        self.assertEqual(record.selection_status, sc.SELECTION_DROPPED)

        select_resp = self.client.post(reverse("select_course", args=[self.stu.sno]), {"cno": self.crs.cno})
        self.assertEqual(select_resp.status_code, 302)
        record.refresh_from_db()
        self.assertEqual(record.selection_status, sc.SELECTION_ACTIVE)
        self.assertEqual(sc.objects.filter(sno=self.stu, cno=self.crs).count(), 1)

    def test_grade_publish_retake_workflow_and_history(self):
        record = sc.objects.create(sno=self.stu, cno=self.crs)
        self.assertTrue(self.client.login(username="teacher_acad", password=DEFAULT_PASSWORD))

        grade_resp = self.client.post(reverse("update_grade", args=[self.stu.sno, self.crs.cno]), {"grade": "88.5"})
        self.assertEqual(grade_resp.status_code, 302)
        record.refresh_from_db()
        self.assertEqual(record.grade, 88.5)
        self.assertEqual(record.grade_status, sc.GRADE_DRAFT)

        publish_resp = self.client.post(reverse("publish_grade", args=[self.stu.sno, self.crs.cno]))
        self.assertEqual(publish_resp.status_code, 302)
        record.refresh_from_db()
        self.assertEqual(record.grade_status, sc.GRADE_PUBLISHED)

        locked_resp = self.client.post(reverse("update_grade", args=[self.stu.sno, self.crs.cno]), {"grade": "90"})
        self.assertEqual(locked_resp.status_code, 302)
        record.refresh_from_db()
        self.assertEqual(record.grade, 88.5)

        retake_resp = self.client.post(reverse("mark_retake", args=[self.stu.sno, self.crs.cno]))
        self.assertEqual(retake_resp.status_code, 302)
        record.refresh_from_db()
        self.assertEqual(record.grade_status, sc.GRADE_RETAKE)
        self.assertIsNone(record.grade)
        self.assertEqual(record.attempt_no, 2)

        actions = list(record.history.values_list("action", flat=True))
        self.assertIn(SelectionHistory.ACTION_GRADE_SAVED, actions)
        self.assertIn(SelectionHistory.ACTION_GRADE_PUBLISHED, actions)
        self.assertIn(SelectionHistory.ACTION_RETAKE_MARKED, actions)

    def test_published_zero_grade_counts_in_official_stats(self):
        record = sc.objects.create(
            sno=self.stu,
            cno=self.crs,
            grade=0,
            selection_status=sc.SELECTION_ACTIVE,
            grade_status=sc.GRADE_PUBLISHED,
        )
        self.assertTrue(self.client.login(username="S3101", password=DEFAULT_PASSWORD))

        course_resp = self.client.get(reverse("student_course", args=[self.stu.sno]))
        detail_resp = self.client.get(reverse("student_detail", args=[self.stu.sno]))
        dashboard_resp = self.client.get(reverse("dashboard"))

        self.assertEqual(course_resp.status_code, 200)
        self.assertContains(course_resp, "0")
        self.assertEqual(detail_resp.context["avg_grade"], 0.0)
        self.assertEqual(dashboard_resp.context["avg_grade"], 0.0)
        record.refresh_from_db()

    def test_published_record_cannot_be_dropped(self):
        record = sc.objects.create(
            sno=self.stu,
            cno=self.crs,
            grade=70,
            selection_status=sc.SELECTION_ACTIVE,
            grade_status=sc.GRADE_PUBLISHED,
        )
        self.assertTrue(self.client.login(username="teacher_acad", password=DEFAULT_PASSWORD))
        resp = self.client.post(reverse("drop_course", args=[self.stu.sno, self.crs.cno]))
        self.assertEqual(resp.status_code, 302)
        record.refresh_from_db()
        self.assertEqual(record.selection_status, sc.SELECTION_ACTIVE)

    def test_teacher_dashboard_is_scoped(self):
        dep2, cls2, crs2 = self.create_school_base(dno="D034", classno="C034", cno="K34")
        other_user = self.create_user("S3200", role="student")
        other_stu = student.objects.create(sno="S3200", user=other_user, sname="学生3200", sex="boy", classno=cls2)
        sc.objects.create(sno=self.stu, cno=self.crs, grade=90, grade_status=sc.GRADE_PUBLISHED)
        sc.objects.create(sno=other_stu, cno=crs2, grade=60, grade_status=sc.GRADE_PUBLISHED)
        self.assertTrue(self.client.login(username="teacher_acad", password=DEFAULT_PASSWORD))

        resp = self.client.get(reverse("dashboard"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["student_total"], 1)
        self.assertEqual(resp.context["course_total"], 1)
        self.assertEqual(resp.context["class_total"], 1)
        self.assertEqual(resp.context["depart_total"], 1)


class MiscAndAuditTests(BaseTestCase):
    def setUp(self):
        self.admin = self.create_user("admin_misc", role="admin")
        self.teacher = self.create_user("teacher_misc", role="teacher")
        self.dep, self.cls, self.crs = self.create_school_base(dno="D051", classno="C051", cno="K51")
        self.assign_teacher(self.teacher, classes=[self.cls], courses=[self.crs])
        self.stu_user = self.create_user("S5101", role="student")
        self.stu = student.objects.create(
            sno="S5101",
            user=self.stu_user,
            sname="学生5101",
            sex="girl",
            classno=self.cls,
        )
        self.record = sc.objects.create(
            sno=self.stu,
            cno=self.crs,
            grade=90,
            selection_status=sc.SELECTION_ACTIVE,
            grade_status=sc.GRADE_PUBLISHED,
        )

    def test_audit_list_admin_only_and_has_grade_workflow_logs(self):
        self.assertTrue(self.client.login(username="teacher_misc", password=DEFAULT_PASSWORD))
        self.client.post(reverse("mark_retake", args=[self.stu.sno, self.crs.cno]))
        self.client.logout()

        self.assertTrue(self.client.login(username="admin_misc", password=DEFAULT_PASSWORD))
        resp = self.client.get(reverse("audit_list"), {"action": "update"})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(AuditLog.objects.filter(model_name="sc").exists())
        self.client.logout()

        self.assertTrue(self.client.login(username="teacher_misc", password=DEFAULT_PASSWORD))
        denied = self.client.get(reverse("audit_list"))
        self.assertEqual(denied.status_code, 302)
        self.assertEqual(denied.url, "/")

    def test_chat_permissions_and_clear(self):
        self.assertTrue(self.client.login(username="S5101", password=DEFAULT_PASSWORD))
        denied = self.client.get(reverse("chat"))
        self.assertEqual(denied.status_code, 302)
        self.assertEqual(denied.url, reverse("dashboard"))
        self.client.logout()

        self.assertTrue(self.client.login(username="teacher_misc", password=DEFAULT_PASSWORD))
        session = self.client.session
        session["chat_messages"] = [
            {"role": "system", "content": "system"},
            {"role": "assistant", "content": "hello"},
            {"role": "user", "content": "keep"},
        ]
        session.save()

        keep_resp = self.client.get(f"{reverse('chat')}?clear=1")
        self.assertEqual(keep_resp.status_code, 200)
        self.assertEqual(len(self.client.session["chat_messages"]), 3)

        clear_resp = self.client.post(reverse("chat"), {"action": "clear"})
        self.assertEqual(clear_resp.status_code, 302)
        self.assertEqual(clear_resp.url, reverse("chat"))
        self.assertEqual(len(self.client.session["chat_messages"]), 2)

    @patch("xx.views_misc.get_ai_response")
    def test_chat_post_with_mocked_ai(self, mocked_ai):
        mocked_ai.return_value = "```python\nresult = list(student.objects.values('sno'))\n```"
        self.assertTrue(self.client.login(username="teacher_misc", password=DEFAULT_PASSWORD))

        resp = self.client.post(reverse("chat"), {"message": "查询学生"})
        self.assertEqual(resp.status_code, 200)
        messages = self.client.session.get("chat_messages", [])
        self.assertGreaterEqual(len(messages), 2)
        self.assertEqual(messages[-1]["role"], "assistant")
