# ============ 标准库 ============
import ast
import builtins
import json
import re
from datetime import datetime, date

import requests
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.password_validation import validate_password
from django.contrib.auth.models import User, Group
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q, Avg, Sum, Count, Max, Min
from django.db.models.query import QuerySet
from django.http import HttpResponse
# ============ Django ============
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.views.generic import ListView, DetailView
# ============ 第三方库 ============
from openpyxl import load_workbook, Workbook

# ============ 本地模块 ============
from .models import student, cl, depart, course, sc, AuditLog
from .audit import log_action, serialize_instance
from .permissions import RoleRequiredMixin, StudentSelfOnlyMixin, is_student, user_has_role

MANAGED_ROLES = ('admin', 'teacher', 'student')


def _get_managed_roles(user):
    if not user or not user.is_authenticated:
        return []
    if user.is_superuser:
        return ['admin']
    return list(user.groups.filter(name__in=MANAGED_ROLES).values_list('name', flat=True))


def _parse_single_role(group_names):
    invalid = [name for name in group_names if name not in MANAGED_ROLES]
    if invalid:
        raise ValueError('检测到非法角色提交')
    if len(group_names) != 1:
        raise ValueError('角色必须且只能选择一个')
    return group_names[0]


def _validate_password_or_raise(password, user=None):
    try:
        validate_password(password, user=user)
    except ValidationError as e:
        raise ValueError('；'.join(e.messages))


# ==================== 用户认证模块 ====================
class UserLoginView(View):
    """用户登录"""
    template_name = 'login.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        if not username or not password:
            messages.error(request, '用户名和密码不能为空')
            return render(request, self.template_name)
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            roles = _get_managed_roles(user)
            if len(roles) > 1:
                messages.error(request, '账号角色配置异常，请联系管理员处理')
                logout(request)
                return redirect('/login/')
            if not roles:
                messages.warning(request, '账号待审核，请联系管理员分配角色')
                logout(request)
                return redirect('/login/')

            role = roles[0]
            if role == 'student':
                if student.objects.filter(sno=user.username).exists():
                    return redirect(f'/students/{user.username}/')
                messages.warning(request, '学生档案未创建，请联系管理员完善信息')
                logout(request)
                return redirect('/login/')
            if role in ('admin', 'teacher'):
                return redirect('/')

            messages.error(request, '账号角色配置异常，请联系管理员处理')
            logout(request)
            return redirect('/login/')
        messages.error(request, '用户名或密码错误')
        return render(request, self.template_name)


class UserLogoutView(LoginRequiredMixin, View):
    """用户登出"""

    def get(self, request):
        logout(request)
        return redirect('/login/')


class UserRegisterView(View):
    template_name = 'register.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        username = request.POST.get('username', '').strip()
        nickname = request.POST.get('nickname', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        if not all([username, nickname, password1, password2]):
            messages.error(request, '所有字段都不能为空')
            return render(request, self.template_name)
        if password1 != password2:
            messages.error(request, '两次密码不一致')
            return render(request, self.template_name)
        try:
            _validate_password_or_raise(password1)
        except ValueError as e:
            messages.error(request, str(e))
            return render(request, self.template_name)
        if User.objects.filter(username=username).exists():
            messages.error(request, '用户名已存在')
            return render(request, self.template_name)

        user = User.objects.create_user(
            username=username,
            password=password1,
            first_name=nickname
        )

        messages.success(request, '注册成功，账号待管理员审核分配角色后方可登录')
        return redirect('/login/')


class UserPasswordView(LoginRequiredMixin, View):
    """修改密码"""
    template_name = 'password.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        old = request.POST.get('old', '')
        new1 = request.POST.get('new1', '')
        new2 = request.POST.get('new2', '')

        if not request.user.check_password(old):
            messages.error(request, '原密码错误')
            return render(request, self.template_name)

        if new1 != new2:
            messages.error(request, '两次新密码不一致')
            return render(request, self.template_name)

        if len(new1) < 6:
            messages.error(request, '密码长度不能少于6位')
            return render(request, self.template_name)
        try:
            _validate_password_or_raise(new1, user=request.user)
        except ValueError as e:
            messages.error(request, str(e))
            return render(request, self.template_name)

        request.user.set_password(new1)
        request.user.save()
        logout(request)
        messages.success(request, '密码修改成功，请重新登录')
        return redirect('/login/')


# ==================== 用户管理模块 ====================

def _user_form_from_request(request):
    return {
        'username': request.POST.get('username', ''),
        'nickname': request.POST.get('nickname', ''),
        'sname': request.POST.get('sname', ''),
        'sex': request.POST.get('sex', ''),
        'classno': request.POST.get('classno', ''),
        'native': request.POST.get('native', ''),
        'age': request.POST.get('age', ''),
        'semester': request.POST.get('semester', ''),
        'home': request.POST.get('home', ''),
        'telephone': request.POST.get('telephone', ''),
        'create_student': request.POST.get('create_student', ''),
    }


def _empty_user_form():
    return {
        'username': '',
        'nickname': '',
        'sname': '',
        'sex': '',
        'classno': '',
        'native': '',
        'age': '',
        'semester': '',
        'home': '',
        'telephone': '',
        'create_student': '',
    }

def _upsert_student_profile_from_request(request, user, existing_profile=None):
    sname = request.POST.get('sname', '').strip()
    classno = request.POST.get('classno', '').strip()

    if not all([sname, classno]):
        raise ValueError('创建学生档案时，姓名和班级不能为空')

    class_obj = cl.objects.get(classno=classno)

    # 验证年龄
    age = request.POST.get('age', '').strip()
    if age:
        try:
            age_int = int(age)
            if age_int < 10 or age_int > 100:
                raise ValueError('年龄必须在10-100之间')
            age = age_int
        except ValueError:
            raise ValueError('年龄必须是有效数字')
    else:
        age = None

    # 验证学期
    semester = request.POST.get('semester', '').strip()
    if semester:
        try:
            semester_int = int(semester)
            if semester_int < 1 or semester_int > 12:
                raise ValueError('学期必须在1-12之间')
            semester = semester_int
        except ValueError:
            raise ValueError('学期必须是有效数字')
    else:
        semester = None

    if existing_profile is None:
        if student.objects.filter(sno=user.username).exists():
            raise ValueError('学生档案已存在')
        existing_profile = student(sno=user.username)

    existing_profile.sname = sname
    existing_profile.sex = request.POST.get('sex', 'girl')
    existing_profile.native = request.POST.get('native', '')
    existing_profile.age = age
    existing_profile.classno = class_obj
    existing_profile.semester = semester
    existing_profile.home = request.POST.get('home', '')
    existing_profile.telephone = request.POST.get('telephone', '')
    existing_profile.save()

class UserListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """用户列表"""
    model = User
    template_name = 'user_list.html'
    context_object_name = 'users'
    paginate_by = 20
    allowed_roles = ('admin',)

    def get_queryset(self):
        queryset = User.objects.all().prefetch_related('groups').order_by('username')

        keyword = self.request.GET.get('q', '').strip()
        role = self.request.GET.get('role', '').strip()

        if keyword:
            queryset = queryset.filter(
                Q(username__icontains=keyword) | Q(first_name__icontains=keyword)
            )
        if role:
            queryset = queryset.filter(groups__name=role).distinct()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        users = list(context['users'])

        usernames = [u.username for u in users]
        student_set = set(
            student.objects.filter(sno__in=usernames).values_list('sno', flat=True)
        )
        context['user_rows'] = [
            {'user': u, 'has_student': u.username in student_set}
            for u in users
        ]
        context['roles'] = Group.objects.order_by('name').values_list('name', flat=True)
        context['q'] = self.request.GET.get('q', '').strip()
        context['role'] = self.request.GET.get('role', '').strip()
        return context


class UserCreateView(LoginRequiredMixin, RoleRequiredMixin, View):
    """创建用户"""
    template_name = 'user_form.html'
    allowed_roles = ('admin',)

    def get(self, request):
        for group_name in MANAGED_ROLES:
            Group.objects.get_or_create(name=group_name)

        return render(request, self.template_name, {
            'user_obj': None,
            'groups': Group.objects.order_by('name'),
            'classes': cl.objects.all(),
            'empty_form': _empty_user_form(),
            'student_profile': None,
            'selected_role': '',
        })

    def post(self, request):
        username = request.POST.get('username', '').strip()
        nickname = request.POST.get('nickname', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        group_names = request.POST.getlist('groups')
        create_student = request.POST.get('create_student') == 'on'

        if not all([username, nickname]):
            messages.error(request, '用户名和昵称不能为空')
            return self._render_with_data(request)

        if password1 or password2:
            if password1 != password2:
                messages.error(request, '两次密码不一致')
                return self._render_with_data(request)

            if len(password1) < 6:
                messages.error(request, '密码长度不能少于6位')
                return self._render_with_data(request)
            try:
                _validate_password_or_raise(password1)
            except ValueError as e:
                messages.error(request, str(e))
                return self._render_with_data(request)
        else:
            messages.error(request, '密码不能为空，请设置密码')
            return self._render_with_data(request)

        try:
            selected_role = _parse_single_role(group_names)
        except ValueError as e:
            messages.error(request, str(e))
            return self._render_with_data(request)

        if User.objects.filter(username=username).exists():
            messages.error(request, '用户名已存在')
            return self._render_with_data(request)

        if selected_role == 'student' and not create_student:
            messages.error(request, '学生角色必须创建学生档案')
            return self._render_with_data(request)

        if create_student and selected_role != 'student':
            messages.error(request, '仅学生角色可以创建学生档案')
            return self._render_with_data(request)

        if create_student:
            if not request.POST.get('sname', '').strip() or not request.POST.get('classno', '').strip():
                messages.error(request, '创建学生档案时，姓名和班级不能为空')
                return self._render_with_data(request)

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username,
                    password=password1,
                    first_name=nickname
                )

                role_group, _ = Group.objects.get_or_create(name=selected_role)
                user.groups.add(role_group)

                if create_student:
                    _upsert_student_profile_from_request(request, user)

            messages.success(request, '用户创建成功')
            return redirect('/users/')

        except cl.DoesNotExist:
            messages.error(request, '班级不存在')
            return self._render_with_data(request)
        except ValueError as e:
            messages.error(request, str(e))
            return self._render_with_data(request)
        except Exception as e:
            messages.error(request, f'创建失败：{str(e)}')
            return self._render_with_data(request)

    def _render_with_data(self, request):
        form_data = _user_form_from_request(request)
        return render(request, self.template_name, {
            'user_obj': None,
            'groups': Group.objects.order_by('name'),
            'classes': cl.objects.all(),
            'form': form_data,
            'empty_form': _empty_user_form(),
            'selected_role': request.POST.getlist('groups')[0] if request.POST.getlist('groups') else '',
            'student_profile': None,
        })

class UserEditView(LoginRequiredMixin, RoleRequiredMixin, View):
    """编辑用户"""
    template_name = 'user_form.html'
    allowed_roles = ('admin',)

    def get(self, request, uid):
        user_obj = get_object_or_404(User, id=uid)
        for group_name in MANAGED_ROLES:
            Group.objects.get_or_create(name=group_name)

        student_profile = student.objects.filter(sno=user_obj.username).first()
        
        return render(request, self.template_name, {
            'user_obj': user_obj,
            'groups': Group.objects.order_by('name'),
            'classes': cl.objects.all(),
            'selected_role': user_obj.groups.filter(name__in=MANAGED_ROLES).values_list('name', flat=True).first() or '',
            'student_profile': student_profile,
            'empty_form': _empty_user_form()
        })

    def post(self, request, uid):
        user_obj = get_object_or_404(User, id=uid)
        nickname = request.POST.get('nickname', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        group_names = request.POST.getlist('groups')
        create_student = request.POST.get('create_student') == 'on'

        if not nickname:
            messages.error(request, '昵称不能为空')
            return self._render_with_data(request, user_obj)

        if password1 or password2:
            if password1 != password2:
                messages.error(request, '两次密码不一致')
                return self._render_with_data(request, user_obj)
            if len(password1) < 6:
                messages.error(request, '密码长度不能少于6位')
                return self._render_with_data(request, user_obj)
            try:
                _validate_password_or_raise(password1, user=user_obj)
            except ValueError as e:
                messages.error(request, str(e))
                return self._render_with_data(request, user_obj)

        try:
            selected_role = _parse_single_role(group_names)
        except ValueError as e:
            messages.error(request, str(e))
            return self._render_with_data(request, user_obj)

        existing_profile = student.objects.filter(sno=user_obj.username).first()

        if selected_role == 'student' and not (create_student or existing_profile):
            messages.error(request, '学生角色必须关联学生档案')
            return self._render_with_data(request, user_obj)

        if create_student and selected_role != 'student':
            messages.error(request, '仅学生角色可以创建学生档案')
            return self._render_with_data(request, user_obj)

        if create_student:
            if not request.POST.get('sname', '').strip() or not request.POST.get('classno', '').strip():
                messages.error(request, '创建学生档案时，姓名和班级不能为空')
                return self._render_with_data(request, user_obj)

        try:
            with transaction.atomic():
                user_obj.first_name = nickname
                if password1:
                    user_obj.set_password(password1)
                user_obj.save()

                user_obj.groups.clear()
                role_group, _ = Group.objects.get_or_create(name=selected_role)
                user_obj.groups.add(role_group)

                if create_student:
                    _upsert_student_profile_from_request(request, user_obj, existing_profile)

            messages.success(request, '用户更新成功')
            return redirect('/users/')

        except cl.DoesNotExist:
            messages.error(request, '班级不存在')
            return self._render_with_data(request, user_obj)
        except ValueError as e:
            messages.error(request, str(e))
            return self._render_with_data(request, user_obj)
        except Exception as e:
            messages.error(request, f'更新失败：{str(e)}')
            return self._render_with_data(request, user_obj)

    def _render_with_data(self, request, user_obj):
        return render(request, self.template_name, {
            'user_obj': user_obj,
            'groups': Group.objects.order_by('name'),
            'classes': cl.objects.all(),
            'form': _user_form_from_request(request),
            'selected_role': request.POST.getlist('groups')[0] if request.POST.getlist('groups') else '',
            'student_profile': student.objects.filter(sno=user_obj.username).first(),
            'empty_form': _empty_user_form()
        })
# ==================== 学生管理模块 ====================

class StudentListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """学生列表"""
    model = student
    template_name = 'student_list.html'
    context_object_name = 'students'
    paginate_by = 10
    allowed_roles = ('admin', 'teacher')

    def get_queryset(self):
        queryset = student.objects.select_related('classno', 'classno__dno')

        # 筛选条件
        sno = self.request.GET.get('sno', '').strip()
        sname = self.request.GET.get('sname', '').strip()
        sex = self.request.GET.get('sex', '').strip()
        classno = self.request.GET.get('classno', '').strip()

        if sno:
            queryset = queryset.filter(sno__icontains=sno)
        if sname:
            queryset = queryset.filter(sname__icontains=sname)
        if sex:
            queryset = queryset.filter(sex=sex)
        if classno:
            queryset = queryset.filter(classno__classno=classno)

        # 排序
        order = self.request.GET.get('order', 'sno')
        direction = self.request.GET.get('direction', 'asc')

        # 排序字段白名单
        order_map = {
            'sno': 'sno',
            'sname': 'sname',
            'age': 'age',
            'classno': 'classno__classno',
            'semester': 'semester',
        }

        order_field = order_map.get(order, 'sno')
        if direction == 'desc':
            order_field = '-' + order_field

        return queryset.order_by(order_field)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()

        context['classes'] = cl.objects.all()
        context['result_count'] = queryset.count()
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_string'] = query_params.urlencode()
        context['boy_count'] = queryset.filter(sex='boy').count()
        context['girl_count'] = queryset.filter(sex='girl').count()
        context['class_count'] = queryset.values('classno').distinct().count()
        context['order'] = self.request.GET.get('order', 'sno')
        context['direction'] = self.request.GET.get('direction', 'asc')

        return context


class StudentAddView(LoginRequiredMixin, RoleRequiredMixin, View):
    """添加学生"""
    template_name = 'student_form.html'
    allowed_roles = ('admin',)

    def get(self, request):
        return render(request, self.template_name, {
            'classes': cl.objects.all()
        })

    def post(self, request):
        try:
            sno = request.POST.get('sno', '').strip()
            sname = request.POST.get('sname', '').strip()
            classno = request.POST.get('classno', '').strip()

            if not all([sno, sname, classno]):
                messages.error(request, '学号、姓名和班级不能为空')
                return render(request, self.template_name, {
                    'classes': cl.objects.all()
                })

            #  检查学号是否已存在
            if student.objects.filter(sno=sno).exists():
                messages.error(request, f'学号 {sno} 已存在')
                return render(request, self.template_name, {
                    'classes': cl.objects.all()
                })

            class_obj = cl.objects.get(classno=classno)

            stu = student.objects.create(
                sno=sno,
                sname=sname,
                sex=request.POST.get('sex', 'girl'),
                native=request.POST.get('native', ''),
                age=request.POST.get('age') or None,
                classno=class_obj,
                semester=request.POST.get('semester') or None,
                home=request.POST.get('home', ''),
                telephone=request.POST.get('telephone', '')
            )
            log_action(
                request,
                'create',
                stu,
                before=None,
                after=serialize_instance(stu)
            )
            messages.success(request, '添加成功')
            return redirect('/students/')

        except cl.DoesNotExist:
            messages.error(request, '班级不存在')
            return render(request, self.template_name, {
                'classes': cl.objects.all()
            })
        except Exception as e:
            messages.error(request, f'添加失败：{str(e)}')
            return render(request, self.template_name, {
                'classes': cl.objects.all()
            })


class StudentEditView(LoginRequiredMixin, RoleRequiredMixin, View):
    """编辑学生"""
    template_name = 'student_form.html'
    allowed_roles = ('admin',)

    def get(self, request, sno):
        stu = get_object_or_404(student, sno=sno)
        return render(request, self.template_name, {
            'stu': stu,
            'classes': cl.objects.all()
        })

    def post(self, request, sno):
        stu = get_object_or_404(student, sno=sno)
        try:
            sname = request.POST.get('sname', '').strip()
            classno = request.POST.get('classno', '').strip()

            if not all([sname, classno]):
                messages.error(request, '姓名和班级不能为空')
                return render(request, self.template_name, {
                    'stu': stu,
                    'classes': cl.objects.all()
                })

            class_obj = cl.objects.get(classno=classno)

            before = serialize_instance(stu)

            stu.sname = sname
            stu.sex = request.POST.get('sex', 'girl')
            stu.native = request.POST.get('native', '')
            stu.age = request.POST.get('age') or None
            stu.classno = class_obj
            stu.semester = request.POST.get('semester') or None
            stu.home = request.POST.get('home', '')
            stu.telephone = request.POST.get('telephone', '')
            stu.save()

            log_action(
                request,
                'update',
                stu,
                before=before,
                after=serialize_instance(stu)
            )

            messages.success(request, '修改成功')
            return redirect('/students/')

        except cl.DoesNotExist:
            messages.error(request, '班级不存在')
            return render(request, self.template_name, {
                'stu': stu,
                'classes': cl.objects.all()
            })
        except Exception as e:
            messages.error(request, f'修改失败：{str(e)}')
            return render(request, self.template_name, {
                'stu': stu,
                'classes': cl.objects.all()
            })


class StudentDeleteView(LoginRequiredMixin, RoleRequiredMixin, View):
    """删除学生"""
    allowed_roles = ('admin',)
    http_method_names = ['post']

    def post(self, request, sno):
        stu = get_object_or_404(student, sno=sno)
        before = serialize_instance(stu)
        stu.delete()
        log_action(
            request,
            'delete',
            stu,
            before=before,
            after=None
        )
        messages.success(request, '删除成功')
        return redirect('/students/')


class StudentDetailView(LoginRequiredMixin, RoleRequiredMixin, StudentSelfOnlyMixin, DetailView):
    """学生详情"""
    model = student
    template_name = 'student_detail.html'
    context_object_name = 'stu'
    pk_url_kwarg = 'sno'
    allowed_roles = ('admin', 'teacher', 'student')

    def get_object(self):
        return get_object_or_404(student, sno=self.kwargs['sno'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        stu = self.get_object()

        records = sc.objects.select_related('cno').filter(sno=stu)
        graded_records = records.filter(grade__isnull=False)

        # ✅ 统一逻辑：学分和平均分都统计已评分课程
        total_credit = graded_records.aggregate(
            total=Sum('cno__credit')
        )['total'] or 0

        avg_grade = graded_records.aggregate(
            avg=Avg('grade')
        )['avg']

        # 及格课程学分（如需要单独统计）
        passed_credit = graded_records.filter(
            grade__gte=60
        ).aggregate(
            total=Sum('cno__credit')
        )['total'] or 0

        context['courses'] = records
        context['total_credit'] = round(total_credit, 1)
        context['passed_credit'] = round(passed_credit, 1)  # ✅ 新增及格学分
        context['avg_grade'] = round(avg_grade, 1) if avg_grade else None
        context['graded_count'] = graded_records.count()

        return context


class StudentImportExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Excel批量导入学生"""
    template_name = 'student_import_excel.html'
    allowed_roles = ('admin',)

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        file = request.FILES.get('file')

        if not file:
            messages.error(request, '请选择 Excel 文件')
            return redirect('/students/import/excel/')

        if not file.name.endswith('.xlsx'):
            messages.error(request, '仅支持 .xlsx 文件')
            return redirect('/students/import/excel/')

        try:
            wb = load_workbook(file)
            ws = wb.active

            headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
            required_headers = [
                'sno', 'sname', 'sex', 'native', 'age',
                'classno', 'semester', 'home', 'telephone'
            ]

            if set(headers) != set(required_headers) or len(headers) != len(required_headers):
                messages.error(request, 'Excel 表头格式不正确，应为：' + ', '.join(required_headers))
                return redirect('/students/import/excel/')

            success = 0
            errors = []

            # ✅ 策略：允许部分成功，每条记录单独事务
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    data = dict(zip(headers, row))

                    # 跳过空行
                    if not data.get('sno'):
                        continue

                    with transaction.atomic():
                        # 检查学号是否已存在
                        if student.objects.filter(sno=data['sno']).exists():
                            raise ValueError(f"学号已存在")

                        try:
                            class_obj = cl.objects.get(classno=data['classno'])
                        except cl.DoesNotExist:
                            raise ValueError(f"班级不存在: {data.get('classno', '')}")

                        student.objects.create(
                            sno=data['sno'],
                            sname=data['sname'],
                            sex=data.get('sex') or 'girl',
                            native=data.get('native') or '',
                            age=data.get('age') or None,
                            classno=class_obj,
                            semester=data.get('semester') or None,
                            home=data.get('home') or '',
                            telephone=data.get('telephone') or '',
                        )
                        success += 1

                except Exception as e:
                    errors.append(f"第{idx}行（学号 {data.get('sno', '未知')}）：{str(e)}")

            if errors:
                error_msg = '；'.join(errors[:5])  # 只显示前5条错误
                if len(errors) > 5:
                    error_msg += f'...（共{len(errors)}条错误）'
                messages.warning(request, f'成功导入 {success} 条，失败 {len(errors)} 条。{error_msg}')
            else:
                messages.success(request, f'成功导入 {success} 条学生')

            return redirect('/students/')

        except Exception as e:
            messages.error(request, f'导入失败：{str(e)}')
            return redirect('/students/import/excel/')


class StudentExportExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    """导出学生Excel"""
    allowed_roles = ('admin', 'teacher')

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = '学生信息'

        headers = [
            'sno', 'sname', 'sex', 'native', 'age',
            'classno', 'semester', 'home', 'telephone'
        ]
        ws.append(headers)

        for stu in student.objects.select_related('classno').all():
            ws.append([
                stu.sno,
                stu.sname,
                stu.sex,
                stu.native or '',
                stu.age or '',
                stu.classno.classno,
                stu.semester or '',
                stu.home or '',
                stu.telephone or ''
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = f'attachment; filename=students_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response


# ==================== 班级管理模块 ====================

class ClassListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """班级列表"""
    model = cl
    template_name = 'class_list.html'
    context_object_name = 'classes'
    allowed_roles = ('admin', 'teacher')

    def get_queryset(self):
        return cl.objects.select_related('dno').annotate(
            student_count=Count('student', distinct=True)
        ).order_by('classno')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student_count'] = student.objects.count()
        context['depart_count'] = depart.objects.count()
        context['class_count'] = self.get_queryset().count()
        return context


class ClassAddView(LoginRequiredMixin, RoleRequiredMixin, View):
    """添加班级"""
    template_name = 'class_form.html'
    allowed_roles = ('admin',)

    def get(self, request):
        return render(request, self.template_name, {
            'departs': depart.objects.all()
        })

    def post(self, request):
        try:
            classno = request.POST.get('classno', '').strip()
            classname = request.POST.get('classname', '').strip()
            dno = request.POST.get('dno', '').strip()

            if not all([classno, classname, dno]):
                messages.error(request, '所有字段不能为空')
                return render(request, self.template_name, {
                    'departs': depart.objects.all()
                })

            if cl.objects.filter(classno=classno).exists():
                messages.error(request, '班级编号已存在')
                return render(request, self.template_name, {
                    'departs': depart.objects.all()
                })

            dno_obj = depart.objects.get(dno=dno)

            c = cl.objects.create(
                classno=classno,
                classname=classname,
                dno=dno_obj
            )
            log_action(
                request,
                'create',
                c,
                before=None,
                after=serialize_instance(c)
            )
            messages.success(request, '添加成功')
            return redirect('/classes/')

        except depart.DoesNotExist:
            messages.error(request, '系部不存在')
            return render(request, self.template_name, {
                'departs': depart.objects.all()
            })
        except Exception as e:
            messages.error(request, f'添加失败：{str(e)}')
            return render(request, self.template_name, {
                'departs': depart.objects.all()
            })


class ClassEditView(LoginRequiredMixin, RoleRequiredMixin, View):
    """编辑班级"""
    template_name = 'class_form.html'
    allowed_roles = ('admin',)

    def get(self, request, classno):
        c = get_object_or_404(cl, classno=classno)
        return render(request, self.template_name, {
            'c': c,
            'departs': depart.objects.all()
        })

    def post(self, request, classno):
        c = get_object_or_404(cl, classno=classno)
        try:
            classname = request.POST.get('classname', '').strip()
            dno = request.POST.get('dno', '').strip()

            if not all([classname, dno]):
                messages.error(request, '所有字段不能为空')
                return render(request, self.template_name, {
                    'c': c,
                    'departs': depart.objects.all()
                })

            dno_obj = depart.objects.get(dno=dno)

            before = serialize_instance(c)

            c.classname = classname
            c.dno = dno_obj
            c.save()

            log_action(
                request,
                'update',
                c,
                before=before,
                after=serialize_instance(c)
            )

            messages.success(request, '修改成功')
            return redirect('/classes/')

        except depart.DoesNotExist:
            messages.error(request, '系部不存在')
            return render(request, self.template_name, {
                'c': c,
                'departs': depart.objects.all()
            })
        except Exception as e:
            messages.error(request, f'修改失败：{str(e)}')
            return render(request, self.template_name, {
                'c': c,
                'departs': depart.objects.all()
            })


class ClassDeleteView(LoginRequiredMixin, RoleRequiredMixin, View):
    """删除班级"""
    allowed_roles = ('admin',)
    http_method_names = ['post']

    def post(self, request, classno):
        c = get_object_or_404(cl, classno=classno)
        before = serialize_instance(c)
        c.delete()
        log_action(
            request,
            'delete',
            c,
            before=before,
            after=None
        )
        messages.success(request, '删除成功')
        return redirect('/classes/')


# ==================== 系部管理模块 ====================

class DepartListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """系部列表"""
    model = depart
    template_name = 'depart_list.html'
    context_object_name = 'departs'
    ordering = ['dno']
    allowed_roles = ('admin', 'teacher')


class DepartAddView(LoginRequiredMixin, RoleRequiredMixin, View):
    """添加系部"""
    template_name = 'depart_form.html'
    allowed_roles = ('admin',)

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        dno = request.POST.get('dno', '').strip()
        dname = request.POST.get('dname', '').strip()
        telephone = request.POST.get('telephone', '').strip()

        if not all([dno, dname]):
            messages.error(request, '系部编号和名称不能为空')
            return render(request, self.template_name)

        if depart.objects.filter(dno=dno).exists():
            messages.error(request, '系部编号已存在')
            return render(request, self.template_name)

        d = depart.objects.create(
            dno=dno,
            dname=dname,
            telephone=telephone
        )
        log_action(
            request,
            'create',
            d,
            before=None,
            after=serialize_instance(d)
        )
        messages.success(request, '添加成功')
        return redirect('/departs/')


class DepartEditView(LoginRequiredMixin, RoleRequiredMixin, View):
    """编辑系部"""
    template_name = 'depart_form.html'
    allowed_roles = ('admin',)

    def get(self, request, dno):
        d = get_object_or_404(depart, dno=dno)
        return render(request, self.template_name, {'d': d})

    def post(self, request, dno):
        d = get_object_or_404(depart, dno=dno)
        dname = request.POST.get('dname', '').strip()
        telephone = request.POST.get('telephone', '').strip()

        if not dname:
            messages.error(request, '系部名称不能为空')
            return render(request, self.template_name, {'d': d})

        before = serialize_instance(d)

        d.dname = dname
        d.telephone = telephone
        d.save()

        log_action(
            request,
            'update',
            d,
            before=before,
            after=serialize_instance(d)
        )

        messages.success(request, '修改成功')
        return redirect('/departs/')


class DepartDeleteView(LoginRequiredMixin, RoleRequiredMixin, View):
    """删除系部"""
    allowed_roles = ('admin',)
    http_method_names = ['post']

    def post(self, request, dno):
        d = get_object_or_404(depart, dno=dno)
        before = serialize_instance(d)
        d.delete()
        log_action(
            request,
            'delete',
            d,
            before=before,
            after=None
        )
        messages.success(request, '删除成功')
        return redirect('/departs/')


# ==================== 课程管理模块 ====================

class CourseListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """课程列表"""
    model = course
    template_name = 'course_list.html'
    context_object_name = 'courses'
    allowed_roles = ('admin', 'teacher')

    def get_queryset(self):
        queryset = course.objects.all()

        cname = self.request.GET.get('cname', '').strip()
        type_ = self.request.GET.get('type', '').strip()
        semester = self.request.GET.get('semester', '').strip()
        order = self.request.GET.get('order', 'cno')
        direction = self.request.GET.get('direction', 'asc')

        if cname:
            queryset = queryset.filter(cname__icontains=cname)
        if type_:
            queryset = queryset.filter(type=type_)
        if semester:
            queryset = queryset.filter(semester=semester)

        # 排序白名单
        allowed_orders = ['cno', 'cname', 'semester', 'credit']
        if order in allowed_orders:
            if direction == 'desc':
                order = '-' + order
            queryset = queryset.order_by(order)

        return queryset


class CourseAddView(LoginRequiredMixin, RoleRequiredMixin, View):
    """添加课程"""
    template_name = 'course_form.html'
    allowed_roles = ('admin',)

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        cno = request.POST.get('cno', '').strip()
        cname = request.POST.get('cname', '').strip()

        if not all([cno, cname]):
            messages.error(request, '课程编号和名称不能为空')
            return render(request, self.template_name)

        if course.objects.filter(cno=cno).exists():
            messages.error(request, '课程编号已存在')
            return render(request, self.template_name)

        c = course.objects.create(
            cno=cno,
            cname=cname,
            lecture=request.POST.get('lecture') or None,
            semester=request.POST.get('semester') or None,
            credit=request.POST.get('credit') or None,
            type=request.POST.get('type', 'crc')
        )
        log_action(
            request,
            'create',
            c,
            before=None,
            after=serialize_instance(c)
        )
        messages.success(request, '添加成功')
        return redirect('/courses/')


class CourseEditView(LoginRequiredMixin, RoleRequiredMixin, View):
    template_name = 'course_form.html'
    allowed_roles = ('admin',)

    def get(self, request, cno):
        c = get_object_or_404(course, cno=cno)
        return render(request, self.template_name, {'c': c})

    def post(self, request, cno):
        c = get_object_or_404(course, cno=cno)
        cname = request.POST.get('cname', '').strip()

        if not cname:
            messages.error(request, '课程名称不能为空')
            return render(request, self.template_name, {'c': c})

        before = serialize_instance(c)

        c.cname = cname
        c.lecture = request.POST.get('lecture') or None
        c.semester = request.POST.get('semester') or None
        c.credit = request.POST.get('credit') or None
        c.type = request.POST.get('type', 'crc')
        c.save()

        log_action(
            request,
            'update',
            c,
            before=before,
            after=serialize_instance(c)
        )

        messages.success(request, '修改成功')
        return redirect('/courses/')


class CourseDeleteView(LoginRequiredMixin, RoleRequiredMixin, View):
    """删除课程"""
    allowed_roles = ('admin',)
    http_method_names = ['post']

    def post(self, request, cno):
        c = get_object_or_404(course, cno=cno)
        before = serialize_instance(c)
        c.delete()
        log_action(
            request,
            'delete',
            c,
            before=before,
            after=None
        )
        messages.success(request, '删除成功')
        return redirect('/courses/')


class SelectCourseView(LoginRequiredMixin, RoleRequiredMixin, View):
    """学生选课"""
    template_name = 'select_course.html'
    allowed_roles = ('admin', 'teacher')

    def get(self, request, sno):
        stu = get_object_or_404(student, sno=sno)

        # ✅ 只显示未选的课程
        selected_courses = sc.objects.filter(sno=stu).values_list('cno_id', flat=True)
        courses = course.objects.exclude(cno__in=selected_courses)

        return render(request, self.template_name, {
            'stu': stu,
            'courses': courses
        })

    def post(self, request, sno):
        stu = get_object_or_404(student, sno=sno)
        cno = request.POST.get('cno', '').strip()

        if not cno:
            messages.error(request, '请选择课程')
            return redirect(f'/select/{sno}/')

        # ✅ 检查是否重复选课
        if sc.objects.filter(sno=stu, cno_id=cno).exists():
            messages.error(request, '已选过该课程')
            return redirect(f'/select/{sno}/')

        try:
            course_obj = course.objects.get(cno=cno)
            sc.objects.create(sno=stu, cno=course_obj)
            messages.success(request, '选课成功')
        except course.DoesNotExist:
            messages.error(request, '课程不存在')
        except Exception as e:
            messages.error(request, f'选课失败：{str(e)}')

        return redirect(f'/sc/{sno}/')


class StudentCourseView(LoginRequiredMixin, RoleRequiredMixin, StudentSelfOnlyMixin, View):
    """学生选课列表"""
    template_name = 'student_course.html'
    allowed_roles = ('admin', 'teacher', 'student')

    def get(self, request, sno):
        stu = get_object_or_404(student, sno=sno)
        records = sc.objects.select_related('cno').filter(sno=stu)

        # ✅ 只统计已评分课程的学分
        graded_records = records.filter(grade__isnull=False)

        total_credit = graded_records.aggregate(
            total=Sum('cno__credit')
        )['total'] or 0

        avg_credit = graded_records.aggregate(
            avg=Avg('cno__credit')
        )['avg'] or 0

        return render(request, self.template_name, {
            'stu': stu,
            'records': records,
            'total_credit': round(total_credit, 1),
            'avg_credit': round(avg_credit, 1),
        })


class UpdateGradeView(LoginRequiredMixin, RoleRequiredMixin, View):
    """录入/修改成绩"""
    template_name = 'grade_form.html'
    allowed_roles = ('admin', 'teacher')

    def get(self, request, sno, cno):
        record = get_object_or_404(sc, sno_id=sno, cno_id=cno)
        return render(request, self.template_name, {'record': record})

    def post(self, request, sno, cno):
        record = get_object_or_404(sc, sno_id=sno, cno_id=cno)
        grade = request.POST.get('grade', '').strip()

        if not grade:
            messages.error(request, '成绩不能为空')
            return render(request, self.template_name, {'record': record})

        try:
            grade_value = float(grade)
            if grade_value < 0 or grade_value > 100:
                messages.error(request, '成绩必须在0-100之间')
                return render(request, self.template_name, {'record': record})

            record.grade = grade_value
            record.save()
            messages.success(request, '成绩录入成功')
            return redirect(f'/sc/{sno}/')

        except ValueError:
            messages.error(request, '成绩必须是数字')
            return render(request, self.template_name, {'record': record})


# ==================== 仪表盘统计模块 ====================

class DashboardView(LoginRequiredMixin, RoleRequiredMixin, View):
    """仪表盘"""
    template_name = 'dashboard.html'
    allowed_roles = ('admin', 'teacher', 'student')

    def get(self, request):
        if is_student(request.user):
            stu = student.objects.filter(sno=request.user.username).select_related('classno').first()
            records = sc.objects.select_related('cno').filter(sno=stu).order_by('-id') if stu else sc.objects.none()
            graded_records = records.filter(grade__isnull=False)

            total_credit = graded_records.aggregate(
                total=Sum('cno__credit')
            )['total'] or 0

            passed_credit = graded_records.filter(
                grade__gte=60
            ).aggregate(
                total=Sum('cno__credit')
            )['total'] or 0

            avg_grade = graded_records.aggregate(
                avg=Avg('grade')
            )['avg']

            return render(request, self.template_name, {
                'stu': stu,
                'recent_sc': records[:10],
                'course_count': records.count(),
                'graded_count': graded_records.count(),
                'avg_grade': round(avg_grade, 1) if avg_grade else None,
                'total_credit': round(total_credit, 1),
                'passed_credit': round(passed_credit, 1),
            })
        # 系部学生人数统计
        depart_stat = student.objects.values(
            'classno__dno__dname'
        ).annotate(
            total=Count('sno')
        ).order_by('-total')

        # 系部课程选课人数统计
        depart_course_stat = sc.objects.values(
            'sno__classno__dno__dname'
        ).annotate(
            total=Count('sno', distinct=True)
        ).order_by('-total')

        # 平均成绩
        avg_grade = sc.objects.filter(
            grade__isnull=False
        ).aggregate(
            avg=Avg('grade')
        )['avg']

        # 最近选课记录
        recent_sc = sc.objects.select_related(
            'sno', 'cno'
        ).order_by('-id')[:10]

        return render(request, self.template_name, {
            'student_total': student.objects.count(),
            'course_total': course.objects.count(),
            'class_total': cl.objects.count(),
            'depart_total': depart.objects.count(),
            'avg_grade': round(avg_grade, 1) if avg_grade else None,
            'depart_stat': depart_stat,
            'depart_course_stat': depart_course_stat,
            'recent_sc': recent_sc,
        })


class CourseStudentsView(LoginRequiredMixin, RoleRequiredMixin, View):
    """课程选课学生列表及成绩统计"""
    template_name = 'course_students.html'
    allowed_roles = ('admin', 'teacher')

    def get(self, request, cno):
        c = get_object_or_404(course, cno=cno)
        records = sc.objects.select_related('sno', 'sno__classno').filter(cno=c)

        # 统计已评分课程
        graded_records = records.filter(grade__isnull=False)

        stats = graded_records.aggregate(
            avg=Avg('grade'),
            max_grade=Max('grade'),
            min_grade=Min('grade'),
            graded=Count('grade')
        )

        # 成绩分布
        excellent = graded_records.filter(grade__gte=90).count()
        good = graded_records.filter(grade__gte=80, grade__lt=90).count()
        passed = graded_records.filter(grade__gte=60, grade__lt=80).count()
        failed = graded_records.filter(grade__lt=60).count()

        return render(request, self.template_name, {
            'course': c,
            'records': records,
            'excellent': excellent,
            'good': good,
            'passed': passed,
            'failed': failed,
            'avg': round(stats['avg'], 1) if stats['avg'] else None,
            'max_grade': stats['max_grade'],
            'min_grade': stats['min_grade'],
            'graded': stats['graded'],
            'total': records.count(),
        })


class AuditLogListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """审计日志"""
    model = AuditLog
    template_name = 'audit_list.html'
    context_object_name = 'logs'
    paginate_by = 20
    allowed_roles = ('admin',)

    def get_queryset(self):
        queryset = AuditLog.objects.select_related('actor').order_by('-created_at')

        action = self.request.GET.get('action', '').strip()
        model_name = self.request.GET.get('model', '').strip()
        actor = self.request.GET.get('actor', '').strip()

        if action:
            queryset = queryset.filter(action=action)
        if model_name:
            queryset = queryset.filter(model_name=model_name)
        if actor:
            queryset = queryset.filter(actor_name__icontains=actor)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_string'] = query_params.urlencode()
        context['model_names'] = (
            AuditLog.objects.order_by('model_name')
            .values_list('model_name', flat=True)
            .distinct()
        )
        context['action_choices'] = AuditLog.ACTION_CHOICES
        return context


# ==================== AI助手模块 ====================

class SecurityError(Exception):
    pass


class CodeValidator:
    """
    校验 AI 生成的代码：只允许“只读 ORM 查询 + 简单 Python 表达式”，禁止任何可能逃逸/破坏的数据操作。
    """

    # 禁止的“直接调用函数名”（Name 调用）
    FORBIDDEN_CALL_NAMES = {
        'eval', 'exec', 'compile', 'open', 'input', 'print',
        '__import__', 'getattr', 'setattr', 'delattr',
        'globals', 'locals', 'vars', 'dir', 'type', 'super',
    }

    # 禁止的“方法名调用”（Attribute 调用）
    FORBIDDEN_METHOD_ATTRS = {
        # ORM 写操作
        'delete', 'update', '_update', '_raw_delete',
        'create', 'bulk_create', 'bulk_update',
        'get_or_create', 'update_or_create',
        'save',

        # 可能绕开 ORM 规则/执行原始 SQL
        'raw', 'extra',

        # 可能引入文件/系统能力（即使没 import）
        'system', 'popen', 'spawn', 'fork',

        # Django/DB 高风险（按需增减）
        'execute',
    }

    @staticmethod
    def validate_ast(code: str) -> bool:
        try:
            tree = ast.parse(code)
        except SyntaxError as e:
            raise SecurityError(f'代码语法错误: {e}')

        for node in ast.walk(tree):
            # 1) 禁止 import / from
            if isinstance(node, (ast.Import, ast.ImportFrom)):
                raise SecurityError('禁止使用 import / from')

            # 2) 禁止 try/except（与你的 prompt 对齐）
            if isinstance(node, ast.Try):
                raise SecurityError('禁止使用 try / except')

            # 3) 禁止定义函数/类/lambda（与你的 prompt 对齐）
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef, ast.Lambda)):
                raise SecurityError('禁止定义函数或类或 lambda')

            # 4) 禁止循环（防死循环卡死 worker）
            if isinstance(node, (ast.For, ast.While)):
                raise SecurityError('禁止使用循环语句')

            # 4.1) 禁止通过下标访问 __builtins__
            if isinstance(node, ast.Subscript):
                if isinstance(node.value, ast.Name) and node.value.id == '__builtins__':
                    raise SecurityError('禁止访问 __builtins__ 下标')

            # 5) 禁止访问任何 dunder 属性（防 __class__/__subclasses__/__mro__ 等逃逸链）
            if isinstance(node, ast.Attribute):
                if isinstance(node.attr, str) and node.attr.startswith('__'):
                    raise SecurityError('禁止访问双下划线属性')

            # 6) 禁止直接引用某些危险名字（尤其是 __builtins__）
            if isinstance(node, ast.Name):
                if node.id in {'__builtins__', '__loader__', '__spec__'}:
                    raise SecurityError(f'禁止使用变量: {node.id}')

            # 7) 禁止危险函数调用 / 危险方法调用
            if isinstance(node, ast.Call):
                # 7.1 Name(...) 调用
                if isinstance(node.func, ast.Name):
                    if node.func.id in CodeValidator.FORBIDDEN_CALL_NAMES:
                        raise SecurityError(f'禁止调用函数: {node.func.id}')

                # 7.2 obj.method(...) 调用
                if isinstance(node.func, ast.Attribute):
                    attr = node.func.attr
                    if isinstance(attr, str) and attr.startswith('_'):
                        raise SecurityError('禁止调用下划线开头方法')
                    if attr in CodeValidator.FORBIDDEN_METHOD_ATTRS:
                        raise SecurityError(f'禁止调用方法: {attr}')
                    # 同时禁止 dunder 方法调用
                    if isinstance(attr, str) and attr.startswith('__'):
                        raise SecurityError('禁止调用双下划线方法')

        return True


class AICodeExecutor:
    """
    执行 AI 生成的代码（只读、可序列化输出），并做安全环境隔离。
    """
    def __init__(self):
        # 只给非常有限的内建函数
        self.safe_builtins = {
            'list', 'dict', 'tuple', 'set',
            'str', 'int', 'float', 'bool',
            'len', 'range', 'enumerate', 'zip',
            'sorted', 'filter', 'map', 'sum',
            'all', 'any', 'min', 'max', 'abs', 'round'
        }

    def execute_ai_code(self, code_string: str, context=None):
        try:
            self._validate_code_safety(code_string)
            exec_globals = self._create_safe_environment()
            if context:
                exec_globals.update(context)

            # 执行 AI 代码（必须产出 result 变量）
            exec(code_string, exec_globals)
            result = exec_globals.get('result')
            return self._serialize_result(result)

        except Exception as e:
            return {'error': f'执行失败: {str(e)}'}

    def _validate_code_safety(self, code: str):
        # 额外做一层快速字符串过滤（AST 才是主防线）
        forbidden_patterns = [
            r'__import__',
            r'open\s*\(',
            r'eval\s*\(',
            r'exec\s*\(',
            r'compile\s*\(',
            r'__builtins__',
        ]
        for pattern in forbidden_patterns:
            if re.search(pattern, code, re.IGNORECASE):
                raise SecurityError(f'检测到不安全代码: {pattern}')

        CodeValidator.validate_ast(code)

    def _create_safe_environment(self):
        # 关键：显式设置 __builtins__，否则 exec 会注入完整 builtins
        safe_builtins_dict = {}
        for name in self.safe_builtins:
            if hasattr(builtins, name):
                safe_builtins_dict[name] = getattr(builtins, name)

        env = {
            '__builtins__': safe_builtins_dict,

            # ORM 可用对象/聚合函数
            'Q': Q, 'Avg': Avg, 'Sum': Sum, 'Count': Count, 'Max': Max, 'Min': Min,
            'student': student,
            'cl': cl,
            'depart': depart,
            'course': course,
            'sc': sc,
        }
        return env

    def _serialize_result(self, result):
        if result is None:
            return {'type': 'none', 'data': '无结果'}

        # multi 表格结构
        if isinstance(result, list) and all(isinstance(r, dict) and 'title' in r and 'data' in r for r in result):
            serialized = []
            for r in result:
                if isinstance(r['data'], QuerySet):
                    r['data'] = list(r['data'][:100])
                r['data'] = make_json_safe(r['data'])
                serialized.append(r)
            return {'type': 'multi', 'data': serialized}

        # queryset
        if isinstance(result, QuerySet):
            data = list(result[:100])
            data = make_json_safe(data)
            return {'type': 'queryset', 'count': result.count(), 'data': data}

        # 常见结构
        if isinstance(result, (list, dict, tuple)):
            return {'type': type(result).__name__, 'data': make_json_safe(result)}

        if isinstance(result, (str, int, float, bool)):
            return {'type': type(result).__name__, 'data': result}

        return {'type': 'other', 'data': str(result)}

CODE_GENERATION_PROMPT = """
你是一个Django ORM代码生成专家。根据用户需求生成可执行的Python代码。
可用的模型：
class depart(models.Model):
    dno = models.CharField(max_length=6, primary_key=True,null=False)
    dname = models.CharField(max_length=10, null=False)
    telephone = models.CharField(max_length=6,)

class cl(models.Model):
    classno = models.CharField(max_length=6,primary_key=True,)
    classname = models.CharField(max_length=10,null=False)
    dno = models.ForeignKey(depart, on_delete=models.CASCADE)
class student(models.Model):
    stusex = (
        ('girl', '女'),
        ('boy', '男'),
    )
    sno = models.CharField(max_length=10, primary_key=True,null=False)
    sname = models.CharField(max_length=10, null=False)
    sex = models.CharField(max_length=4,choices=stusex, default='girl')
    native = models.CharField(max_length=20,)
    age = models.IntegerField(null=True)
    classno = models.ForeignKey(cl, on_delete=models.CASCADE)
    entime = models.DateTimeField(null=True,auto_now=True)
    semester = models.IntegerField(null=True)
    home = models.CharField(max_length=40,)
    telephone = models.CharField(max_length=20, )
class course(models.Model):
    coutype = (
        ('crc', '公共课'),
        ('bcim', '专业基础课'),
        ('spc', '专业课'),
        ('ocos', '选修课')
    )
    cno = models.CharField(max_length=3, primary_key=True,null=False)
    cname = models.CharField(max_length=20, null=False)
    lecture = models.FloatField(null=True)
    semester = models.IntegerField(null=True)
    credit = models.FloatField(null=True)
    type = models.CharField(max_length=10,null = True,choices=coutype,default='crc')
class sc(models.Model):
    sno = models.ForeignKey(student, on_delete=models.CASCADE)
    cno = models.ForeignKey(course, on_delete=models.CASCADE)
    grade = models.FloatField(null=True)
生成要求：
如果返回多个模型的数据，请使用列表，每个元素包含 title 和 data
严格按照上面给出的模型以及字段名来进行编写代码，不允许假设，不允许更改。
1. 只使用Django ORM查询，不要使用原始SQL
2. 查询结果必须赋值给变量 `result`
3. 代码必须安全，不能包含文件操作、系统调用等
4. 优先使用values()获取字典格式数据
5. 不需要异常处理，直接写查询并赋值 result
6. 不允许出现 import / from / print / try / except
7. 不允许定义函数或类
8. 可以直接使用：student, cl, depart, course, sc, Q, Count, Avg, Sum
10.你可以使用跨表的多表查询
11.course 模型 type 字段合法取值：
- "crc" → 公共课
- "bcim" → 专业基础课
- "spc" → 专业课
- "ocos" → 选修课
前者是具体的值，后者是前者的含义
示例：
用户：查询所有男生信息
代码：
result = student.objects.filter(sex='boy').values('sno', 'sname', 'age')
用户：统计每个班级的学生人数
代码：
result = list(student.objects.values('classno__classname').annotate(count=Count('sno')))

现在请为以下需求生成代码：
用户需求：{user_query}
"""


def get_ai_response(messages):
    url = f"{settings.AI_BASE_URL}/v1/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {settings.AI_API_KEY}",
    }
    payload = {
        "model": settings.AI_MODEL,  # 比如 'deepseek-chat'
        "messages": messages
    }

    try:
        response = requests.post(
            url,
            headers=headers,
            json=payload,
            timeout=30,
        )

        if response.status_code == 401:
            raise RuntimeError(
                "AI接口返回 401 未认证：请检查 DeepSeek API Key 是否配置正确、是否带在 Authorization 头里。"
            )

        if not response.ok:
            raise RuntimeError(
                f"AI接口返回非成功状态码 {response.status_code}，内容: {response.text[:200]}"
            )

        data = response.json()
        return data["choices"][0]["message"]["content"]

    except Exception as e:
        raise RuntimeError(f"AI调用失败: {str(e)}")


def extract_code_from_response(text: str) -> str:
    """
    优先提取 ```...``` 内的代码；兼容 ```python / ```py / ``` 以及 \r\n。
    若没有代码块，按原逻辑兜底截取。
    """
    if not text:
        return ""

    code_match = re.search(r'```(?:python|py)?\s*\r?\n(.*?)\r?\n```', text, re.DOTALL | re.IGNORECASE)
    if code_match:
        return code_match.group(1).strip()

    lines = text.splitlines()
    code_lines = []
    in_code = False

    for line in lines:
        if line.strip().startswith("```") and in_code:
            break

        if any(k in line for k in ['result =', 'result=', 'def ', 'class ', 'import ', 'from ']):
            in_code = True

        if in_code:
            if re.match(r'^\s*(解释|说明|注意|结果|输出|AI回复|以下是)\s*[:：]?\s*$', line.strip()):
                break
            if line.strip() and not line.lstrip().startswith('#'):
                code_lines.append(line)

    return '\n'.join(code_lines).strip() if code_lines else text.strip()


def format_execution_result(result):
    if 'error' in result:
        return json.dumps([{"error": result['error']}], ensure_ascii=False)
    if result.get('type') == 'multi' and isinstance(result.get('data'), list):
        formatted_data = []
        for tbl in result['data']:
            if isinstance(tbl, dict) and 'title' in tbl and 'data' in tbl:
                formatted_data.append({
                    'title': tbl['title'],
                    'data': tbl['data']
                })
        return json.dumps([{
            'type': 'multi',
            'data': formatted_data
        }], ensure_ascii=False)
    if result.get('type') == 'queryset' and isinstance(result.get('data'), list):
        return json.dumps([{
            'type': 'queryset',
            'count': result.get('count', len(result['data'])),
            'data': result['data']
        }], ensure_ascii=False)
    if result.get('type') in ('list', 'dict', 'tuple'):
        return json.dumps([{
            'type': result['type'],
            'data': result['data']
        }], ensure_ascii=False)
    if result.get('type') in ('str', 'int', 'float', 'bool'):
        return json.dumps([{
            'type': result['type'],
            'data': result['data']
        }], ensure_ascii=False)
    return json.dumps([{
        'type': 'other',
        'data': str(result.get('data'))
    }], ensure_ascii=False)


def make_json_safe(obj):
    if isinstance(obj, dict):
        return {k: make_json_safe(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [make_json_safe(v) for v in obj]
    elif isinstance(obj, tuple):
        return tuple(make_json_safe(v) for v in obj)
    elif isinstance(obj, (datetime, date)):
        return obj.strftime('%Y-%m-%d %H:%M:%S')
    elif hasattr(obj, '__dict__'):
        return {k: make_json_safe(v) for k, v in obj.__dict__.items() if not k.startswith('_')}
    else:
        return obj


@login_required
def chat_view(request):
    if is_student(request.user):
        messages.error(request, '无权限访问')
        return redirect('/')
    if not request.user.groups.filter(name__in=['admin', 'teacher']).exists() and not request.user.is_superuser:
        messages.error(request, '无权限访问')
        return redirect('/')
    if request.GET.get("clear") == "1":
        request.session["chat_messages"] = [
            {
                "role": "system",
                "content": """你是一个Django ORM代码生成助手。根据用户需求生成可直接执行的Python代码。
    代码应该简洁、安全，并且将结果赋值给变量`result`。
    可用的模型：student, cl, depart, course, sc。
    使用Django ORM进行查询，不要使用原始SQL。"""
            },
            {
                "role": "assistant",
                "content": "你好我是你的AI助手，我可以帮助你完成查询工作！"
            }
        ]
        request.session.modified = True
        return redirect("/chat/")
    if "chat_messages" not in request.session:
        request.session["chat_messages"] = [{
            "role": "system",
            "content": """你是一个Django ORM代码生成助手。根据用户需求生成可直接执行的Python代码。
代码应该简洁、安全，并且将结果赋值给变量`result`。
可用的模型：student, cl, depart, course, sc。
使用Django ORM进行查询，不要使用原始SQL。"""
        }]
    executor = AICodeExecutor()
    if request.method == "POST":
        user_input = request.POST.get("message", "").strip()
        if not user_input:
            return render(request, "chat.html", {"messages": request.session["chat_messages"]})
        request.session["chat_messages"].append({"role": "user", "content": user_input})
        try:
            prompt = CODE_GENERATION_PROMPT.format(user_query=user_input)
            ai_response = get_ai_response([
                {"role": "system", "content": prompt},
                {"role": "user", "content": user_input}
            ])
            code = extract_code_from_response(ai_response)
            if 'cno__cname' in code:
                code = code.replace(
                    'values(\'cno__cname\')',
                    'values_list(\'cno__cname\', flat=True)'
                )
            execution_result = executor.execute_ai_code(code)
            if 'error' in execution_result:
                reply = f"执行错误:\n{execution_result['error']}\n\n生成的代码:\n```python\n{code}\n```"
            else:
                reply = format_execution_result(execution_result)
        except Exception as e:
            reply = f"处理失败:\n{str(e)}\n\nAI回复:\n{ai_response if 'ai_response' in locals() else '无'}"
        request.session["chat_messages"].append({"role": "assistant", "content": reply})
        request.session.modified = True
    return render(request, "chat.html", {"messages": request.session["chat_messages"]})
