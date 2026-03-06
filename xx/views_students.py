from datetime import datetime

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import Group
from django.db import transaction
from django.db.models import Avg, F, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views import View
from django.views.generic import DetailView, ListView
from openpyxl import Workbook, load_workbook

from .audit import log_action, serialize_instance
from .forms import StudentAddForm, StudentEditForm
from .models import StudentImportJob, UserAccount, cl, sc, student
from .permissions import (
    RoleRequiredMixin,
    StudentSelfOnlyMixin,
    filter_students_for_user,
    is_teacher,
    teacher_can_access_student,
    teacher_class_queryset,
)
from .view_shared import (
    DEFAULT_STUDENT_PASSWORD,
    ensure_student_user_account,
    flash_form_errors,
    set_account_status,
)


IMPORT_HEADERS = [
    'sno', 'sname', 'sex', 'native', 'age',
    'classno', 'semester', 'home', 'telephone',
]


def _official_records(queryset):
    return queryset.filter(
        selection_status=sc.SELECTION_ACTIVE,
        grade_status=sc.GRADE_PUBLISHED,
        grade__isnull=False,
    )


def _active_student_or_404(sno):
    return get_object_or_404(student.objects.select_related('classno', 'classno__dno'), sno=sno, is_active=True)


def _serialize_import_row(row_number, cleaned, operation):
    return {
        'row_number': row_number,
        'operation': operation,
        'sno': cleaned['sno'].strip(),
        'sname': cleaned['sname'].strip(),
        'sex': cleaned.get('sex') or 'girl',
        'native': cleaned.get('native') or '',
        'age': cleaned.get('age'),
        'classno': cleaned['classno'].classno,
        'semester': cleaned.get('semester'),
        'home': cleaned.get('home') or '',
        'telephone': cleaned.get('telephone') or '',
    }


def _apply_student_row(payload):
    sno = payload['sno']
    existing = student.objects.filter(sno=sno).first()
    matched_user, account_created = ensure_student_user_account(sno, payload['sname'])
    if student.objects.filter(user=matched_user, is_active=True).exclude(sno=sno).exists():
        raise ValueError('账号已绑定其他有效学生档案')

    if existing is None:
        before = None
        stu = student(sno=sno)
        action = 'create'
    else:
        before = serialize_instance(existing)
        stu = existing
        action = 'update'

    stu.user = matched_user
    stu.sname = payload['sname']
    stu.sex = payload['sex']
    stu.native = payload['native']
    stu.age = payload['age']
    stu.classno_id = payload['classno']
    stu.semester = payload['semester']
    stu.home = payload['home']
    stu.telephone = payload['telephone']
    stu.restore()
    stu.save()
    return stu, action, before, account_created


class StudentListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    model = student
    template_name = 'student_list.html'
    context_object_name = 'students'
    paginate_by = 10
    allowed_roles = ('admin', 'teacher')

    def get_queryset(self):
        scope = self.request.GET.get('scope', 'active').strip()
        queryset = student.objects.select_related('classno', 'classno__dno')
        if scope == 'active':
            queryset = queryset.filter(is_active=True)
        elif scope == 'inactive':
            queryset = queryset.filter(is_active=False)

        if is_teacher(self.request.user):
            queryset = filter_students_for_user(self.request.user, queryset)

        sno = self.request.GET.get('sno', '').strip()
        sname = self.request.GET.get('sname', '').strip()
        sex = self.request.GET.get('sex', '').strip()
        classno = self.request.GET.get('classno', '').strip()

        if sno:
            queryset = queryset.filter(sno__icontains=sno)
        if sname:
            queryset = queryset.filter(sname__icontains=sname)
        if sex:
            queryset = queryset.filter(sex=sex)
        if classno:
            queryset = queryset.filter(classno__classno=classno)

        order = self.request.GET.get('order', 'sno')
        direction = self.request.GET.get('direction', 'asc')
        order_map = {
            'sno': 'sno',
            'sname': 'sname',
            'age': 'age',
            'classno': 'classno__classno',
            'semester': 'semester',
        }

        order_field = order_map.get(order, 'sno')
        ordering = [F(order_field).desc(nulls_last=True)] if direction == 'desc' else [F(order_field).asc(nulls_last=True)]
        if order_field != 'sno':
            secondary = F('sno').desc() if direction == 'desc' else F('sno').asc()
            ordering.append(secondary)

        return queryset.order_by(*ordering)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        classes = teacher_class_queryset(self.request.user) if is_teacher(self.request.user) else cl.objects.filter(is_active=True).order_by('classno')

        context['classes'] = classes
        context['result_count'] = queryset.count()
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_string'] = query_params.urlencode()
        context['boy_count'] = queryset.filter(sex='boy').count()
        context['girl_count'] = queryset.filter(sex='girl').count()
        context['class_count'] = queryset.values('classno').distinct().count()
        context['order'] = self.request.GET.get('order', 'sno')
        context['direction'] = self.request.GET.get('direction', 'asc')
        context['scope'] = self.request.GET.get('scope', 'active').strip()
        return context


class StudentAddView(LoginRequiredMixin, RoleRequiredMixin, View):
    template_name = 'student_form.html'
    allowed_roles = ('admin',)

    def get(self, request):
        return render(request, self.template_name, {'classes': cl.objects.filter(is_active=True).order_by('classno')})

    def post(self, request):
        form = StudentAddForm(request.POST)
        if not form.is_valid():
            flash_form_errors(request, form)
            return render(request, self.template_name, {'classes': cl.objects.filter(is_active=True).order_by('classno')})

        try:
            sno = form.cleaned_data['sno'].strip()
            sname = form.cleaned_data['sname'].strip()
            class_obj = form.cleaned_data['classno']
            existing = student.objects.filter(sno=sno).first()

            if existing and existing.is_active:
                messages.error(request, f'学号 {sno} 已存在')
                return render(request, self.template_name, {'classes': cl.objects.filter(is_active=True).order_by('classno')})

            matched_user, account_created = ensure_student_user_account(sno, sname)
            if student.objects.filter(user=matched_user, is_active=True).exclude(sno=sno).exists():
                messages.error(request, f'账号 {sno} 已绑定其他有效学生档案')
                return render(request, self.template_name, {'classes': cl.objects.filter(is_active=True).order_by('classno')})

            if existing:
                before = serialize_instance(existing)
                stu = existing
                action = 'update'
            else:
                before = None
                stu = student(sno=sno)
                action = 'create'

            stu.user = matched_user
            stu.sname = sname
            stu.sex = form.cleaned_data.get('sex') or 'girl'
            stu.native = form.cleaned_data.get('native') or ''
            stu.age = form.cleaned_data.get('age') or None
            stu.classno = class_obj
            stu.semester = form.cleaned_data.get('semester') or None
            stu.home = form.cleaned_data.get('home') or ''
            stu.telephone = form.cleaned_data.get('telephone') or ''
            stu.restore()
            stu.save()

            log_action(request, action, stu, before=before, after=serialize_instance(stu))
            if existing:
                messages.success(request, '学生档案已恢复并更新')
            elif account_created:
                messages.success(request, f'添加成功，学生初始密码为 {DEFAULT_STUDENT_PASSWORD}')
            else:
                messages.success(request, '添加成功')
            return redirect('student_list')
        except Exception as exc:
            messages.error(request, f'添加失败：{str(exc)}')
            return render(request, self.template_name, {'classes': cl.objects.filter(is_active=True).order_by('classno')})


class StudentEditView(LoginRequiredMixin, RoleRequiredMixin, View):
    template_name = 'student_form.html'
    allowed_roles = ('admin',)

    def get(self, request, sno):
        stu = _active_student_or_404(sno)
        return render(request, self.template_name, {'stu': stu, 'classes': cl.objects.filter(is_active=True).order_by('classno')})

    def post(self, request, sno):
        stu = _active_student_or_404(sno)
        form = StudentEditForm(request.POST)
        if not form.is_valid():
            flash_form_errors(request, form)
            return render(request, self.template_name, {'stu': stu, 'classes': cl.objects.filter(is_active=True).order_by('classno')})

        try:
            before = serialize_instance(stu)
            stu.sname = form.cleaned_data['sname'].strip()
            stu.sex = form.cleaned_data.get('sex') or 'girl'
            stu.native = form.cleaned_data.get('native') or ''
            stu.age = form.cleaned_data.get('age') or None
            stu.classno = form.cleaned_data['classno']
            stu.semester = form.cleaned_data.get('semester') or None
            stu.home = form.cleaned_data.get('home') or ''
            stu.telephone = form.cleaned_data.get('telephone') or ''
            stu.save()

            log_action(request, 'update', stu, before=before, after=serialize_instance(stu))
            messages.success(request, '修改成功')
            return redirect('student_list')
        except Exception as exc:
            messages.error(request, f'修改失败：{str(exc)}')
            return render(request, self.template_name, {'stu': stu, 'classes': cl.objects.filter(is_active=True).order_by('classno')})


class StudentDeleteView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ('admin',)
    http_method_names = ['post']

    def post(self, request, sno):
        stu = _active_student_or_404(sno)
        account_action = request.POST.get('account_action', 'disable')
        if account_action not in {'disable', 'pending', 'unlink'}:
            account_action = 'disable'

        before = serialize_instance(stu)
        bound_user = stu.user
        stu.archive()
        if account_action == 'unlink':
            stu.user = None
        stu.save()

        if bound_user:
            student_group = Group.objects.filter(name='student').first()
            if student_group:
                bound_user.groups.remove(student_group)
            status = UserAccount.STATUS_DISABLED if account_action == 'disable' else UserAccount.STATUS_PENDING
            set_account_status(bound_user, status, reviewer=request.user, notes='student_archived')

        log_action(request, 'delete', stu, before=before, after=serialize_instance(stu))
        messages.success(request, '学生档案已归档')
        return redirect('student_list')


class StudentDetailView(LoginRequiredMixin, RoleRequiredMixin, StudentSelfOnlyMixin, DetailView):
    model = student
    template_name = 'student_detail.html'
    context_object_name = 'stu'
    pk_url_kwarg = 'sno'
    allowed_roles = ('admin', 'teacher', 'student')

    def dispatch(self, request, *args, **kwargs):
        if is_teacher(request.user):
            stu = student.objects.filter(sno=kwargs['sno'], is_active=True).select_related('classno').first()
            if stu is None or not teacher_can_access_student(request.user, stu):
                messages.error(request, '无权限访问')
                return redirect('/')
        return super().dispatch(request, *args, **kwargs)

    def get_object(self):
        return _active_student_or_404(self.kwargs['sno'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        stu = self.get_object()

        records = sc.objects.select_related('cno').filter(sno=stu).order_by('cno__cno')
        official_records = _official_records(records)
        total_credit = official_records.aggregate(total=Sum('cno__credit'))['total'] or 0
        avg_grade = official_records.aggregate(avg=Avg('grade'))['avg']
        passed_credit = official_records.filter(grade__gte=60).aggregate(total=Sum('cno__credit'))['total'] or 0

        context['courses'] = records
        context['total_credit'] = round(total_credit, 1)
        context['passed_credit'] = round(passed_credit, 1)
        context['avg_grade'] = round(avg_grade, 1) if avg_grade is not None else None
        context['graded_count'] = official_records.count()
        return context


class StudentImportExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    template_name = 'student_import_excel.html'
    allowed_roles = ('admin',)

    def get(self, request):
        job = None
        job_id = request.GET.get('job_id')
        if job_id:
            job = StudentImportJob.objects.filter(id=job_id).first()
        return render(request, self.template_name, {'job': job})

    def post(self, request):
        action = request.POST.get('action', 'preview')
        if action == 'apply':
            return self._apply_job(request)
        return self._preview_job(request)

    def _preview_job(self, request):
        file = request.FILES.get('file')
        mode = request.POST.get('mode', StudentImportJob.MODE_CREATE)

        if mode not in {StudentImportJob.MODE_CREATE, StudentImportJob.MODE_UPDATE}:
            mode = StudentImportJob.MODE_CREATE
        if not file:
            messages.error(request, '请选择 Excel 文件')
            return redirect('student_import_excel')
        if not file.name.endswith('.xlsx'):
            messages.error(request, '仅支持 .xlsx 文件')
            return redirect('student_import_excel')

        try:
            wb = load_workbook(file)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
            if headers != IMPORT_HEADERS:
                messages.error(request, 'Excel 表头格式不正确，应为：' + ', '.join(IMPORT_HEADERS))
                return redirect('student_import_excel')

            preview_rows = []
            error_rows = []
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                data = dict(zip(headers, row))
                if not data.get('sno'):
                    continue

                row_form = StudentAddForm({
                    'sno': str(data.get('sno') or '').strip(),
                    'sname': str(data.get('sname') or '').strip(),
                    'sex': str(data.get('sex') or '').strip(),
                    'native': str(data.get('native') or '').strip(),
                    'age': data.get('age'),
                    'classno': str(data.get('classno') or '').strip(),
                    'semester': data.get('semester'),
                    'home': str(data.get('home') or '').strip(),
                    'telephone': str(data.get('telephone') or '').strip(),
                })
                if not row_form.is_valid():
                    error_items = []
                    for field, field_errors in row_form.errors.items():
                        error_items.append(f'{field}: {"；".join(field_errors)}')
                    error_rows.append({
                        'row_number': idx,
                        'sno': str(data.get('sno') or '').strip(),
                        'error': '；'.join(error_items),
                    })
                    continue

                cleaned = row_form.cleaned_data
                existing = student.objects.filter(sno=cleaned['sno']).first()
                if mode == StudentImportJob.MODE_CREATE:
                    if existing and existing.is_active:
                        error_rows.append({
                            'row_number': idx,
                            'sno': cleaned['sno'],
                            'error': '学号已存在，新增模式不允许覆盖',
                        })
                        continue
                    operation = 'restore' if existing else 'create'
                else:
                    if existing is None:
                        error_rows.append({
                            'row_number': idx,
                            'sno': cleaned['sno'],
                            'error': '更新模式要求学号已存在',
                        })
                        continue
                    operation = 'restore_update' if not existing.is_active else 'update'

                preview_rows.append(_serialize_import_row(idx, cleaned, operation))

            summary = {
                'preview_count': len(preview_rows),
                'error_count': len(error_rows),
            }
            job = StudentImportJob.objects.create(
                creator=request.user,
                mode=mode,
                status=StudentImportJob.STATUS_PREVIEWED,
                preview_rows=preview_rows,
                error_rows=error_rows,
                summary=summary,
            )
            return render(request, self.template_name, {'job': job})
        except Exception as exc:
            messages.error(request, f'预检失败：{str(exc)}')
            return redirect('student_import_excel')

    def _apply_job(self, request):
        job_id = request.POST.get('job_id')
        job = get_object_or_404(StudentImportJob, id=job_id)
        if job.status == StudentImportJob.STATUS_APPLIED:
            messages.warning(request, '该导入任务已执行过')
            return redirect(f"{reverse('student_import_excel')}?job_id={job.id}")

        created = 0
        updated = 0
        restored = 0
        created_accounts = 0
        errors = []

        for row in job.preview_rows:
            try:
                with transaction.atomic():
                    existing = student.objects.filter(sno=row['sno']).first()
                    was_existing = existing is not None
                    was_inactive = bool(existing and not existing.is_active)
                    stu, action, before, account_created = _apply_student_row(row)
                    log_action(request, action, stu, before=before, after=serialize_instance(stu))
                    if not was_existing:
                        created += 1
                    elif was_inactive:
                        restored += 1
                    else:
                        updated += 1
                    if account_created:
                        created_accounts += 1
            except Exception as exc:
                errors.append(f"第{row['row_number']}行（学号 {row['sno']}）：{str(exc)}")

        job.status = StudentImportJob.STATUS_APPLIED
        job.summary = {
            'created': created,
            'updated': updated,
            'restored': restored,
            'created_accounts': created_accounts,
            'error_count': len(errors),
        }
        job.applied_at = timezone.now()
        job.save(update_fields=['status', 'summary', 'applied_at', 'updated_at'])

        if errors:
            messages.warning(request, f'导入完成：新增 {created}，更新 {updated}，恢复 {restored}，失败 {len(errors)}')
        else:
            messages.success(request, f'导入完成：新增 {created}，更新 {updated}，恢复 {restored}')
        return redirect(f"{reverse('student_import_excel')}?job_id={job.id}")


class StudentExportExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ('admin', 'teacher')

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = '学生信息'
        headers = ['sno', 'sname', 'sex', 'native', 'age', 'classno', 'semester', 'home', 'telephone']
        ws.append(headers)

        queryset = student.objects.select_related('classno').filter(is_active=True).order_by('sno')
        if is_teacher(request.user):
            queryset = filter_students_for_user(request.user, queryset)

        for stu in queryset:
            ws.append([
                stu.sno,
                stu.sname,
                stu.sex,
                stu.native or '',
                stu.age or '',
                stu.classno.classno,
                stu.semester or '',
                stu.home or '',
                stu.telephone or '',
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename=students_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        wb.save(response)
        return response
